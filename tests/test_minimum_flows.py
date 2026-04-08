import os
import sqlite3
import threading
import unittest
from unittest.mock import patch
from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO
from pathlib import Path

os.environ["DATABASE_URL"] = "sqlite+pysqlite:///./test_minimum_flows.db"
os.environ.setdefault("SECRET_KEY", "test-secret-key-at-least-16")
os.environ.setdefault("ALGORITHM", "HS256")
os.environ.setdefault("ACCESS_TOKEN_EXPIRE_MINUTES", "60")
os.environ.setdefault("EXCEPTION_WORKFLOW_SYSTEM_KEY", "test-exception-system-key")

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from fastapi import HTTPException
from sqlalchemy import event

from app.core.db import Base, SessionLocal, engine
from app.core.security import create_access_token, hash_password, hash_token
from app.main import app
from app.api import attendance as attendance_api
from app.models import AttendanceException, AttendanceExceptionAudit, AttendanceExceptionNotification, AttendanceLog, CheckinRule, Employee, Group, GroupGeofence, PasswordResetToken, RefreshToken, User
from app.schemas.attendance import LocationRequest
from app.services.auth.password_reset_service import PasswordResetService, cleanup_password_reset_tokens
from app.services.attendance_exception_jobs import expire_overdue_exceptions, send_expire_reminders


class _BoolOr:
    def __init__(self) -> None:
        self.value = False

    def step(self, item) -> None:
        if item:
            self.value = True

    def finalize(self) -> int:
        return 1 if self.value else 0


class _FixedDateTime(datetime):
    fixed_now: datetime | None = None

    @classmethod
    def now(cls, tz=None):
        if cls.fixed_now is None:
            return super().now(tz=tz)
        if tz is None:
            return cls.fixed_now.replace(tzinfo=None)
        return cls.fixed_now.astimezone(tz)

class _SpyMailSender:
    def __init__(self) -> None:
        self.sent = []

    def send_reset_password(self, payload) -> None:
        self.sent.append(payload)

    def send_exception_notification(self, payload) -> None:
        self.sent.append(payload)


@event.listens_for(engine, "connect")
def _register_sqlite_bool_or(dbapi_connection, _connection_record) -> None:
    if isinstance(dbapi_connection, sqlite3.Connection):
        dbapi_connection.create_aggregate("bool_or", 1, _BoolOr)


class MinimumFlowsTestCase(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.db_path = Path("test_minimum_flows.db")
        if cls.db_path.exists():
            cls.db_path.unlink()

        engine.dispose()
        Base.metadata.create_all(bind=engine)
        cls.client = TestClient(app)

    @classmethod
    def tearDownClass(cls) -> None:
        cls.client.close()
        engine.dispose()
        if cls.db_path.exists():
            cls.db_path.unlink()

    def setUp(self) -> None:
        with SessionLocal() as db:
            db.query(AttendanceExceptionNotification).delete()
            db.query(AttendanceExceptionAudit).delete()
            db.query(AttendanceException).delete()
            db.query(AttendanceLog).delete()
            db.query(RefreshToken).delete()
            db.query(PasswordResetToken).delete()
            db.query(Employee).delete()
            db.query(GroupGeofence).delete()
            db.query(Group).delete()
            db.query(CheckinRule).delete()
            db.query(User).delete()
            db.commit()

    def _create_user(self, email: str, password: str, role: str) -> User:
        with SessionLocal() as db:
            user = User(email=email, password_hash=hash_password(password), role=role)
            db.add(user)
            db.commit()
            db.refresh(user)
            return user

    def _create_group(
        self,
        code: str,
        name: str,
        active: bool = True,
        start_time: time | None = None,
        grace_minutes: int | None = None,
        end_time: time | None = None,
        checkout_grace_minutes: int | None = None,
        cross_day_cutoff_minutes: int | None = None,
    ) -> Group:
        with SessionLocal() as db:
            group = Group(
                code=code,
                name=name,
                active=active,
                start_time=start_time,
                grace_minutes=grace_minutes,
                end_time=end_time,
                checkout_grace_minutes=checkout_grace_minutes,
                cross_day_cutoff_minutes=cross_day_cutoff_minutes,
            )
            db.add(group)
            db.commit()
            db.refresh(group)
            return group

    def _create_geofence(
        self,
        group_id: int,
        name: str,
        latitude: float,
        longitude: float,
        radius_m: int,
        active: bool = True,
    ) -> GroupGeofence:
        with SessionLocal() as db:
            geofence = GroupGeofence(
                group_id=group_id,
                name=name,
                latitude=latitude,
                longitude=longitude,
                radius_m=radius_m,
                active=active,
            )
            db.add(geofence)
            db.commit()
            db.refresh(geofence)
            return geofence

    def _create_employee(
        self,
        code: str,
        full_name: str,
        user_id: int | None,
        group_id: int | None = None,
    ) -> Employee:
        with SessionLocal() as db:
            emp = Employee(code=code, full_name=full_name, user_id=user_id, group_id=group_id)
            db.add(emp)
            db.commit()
            db.refresh(emp)
            return emp

    def _create_rule(
        self,
        latitude: float = 10.7769,
        longitude: float = 106.7009,
        radius_m: int = 200,
        start_time: time = time(8, 0),
        grace_minutes: int = 30,
        end_time: time = time(17, 30),
        checkout_grace_minutes: int = 0,
        cross_day_cutoff_minutes: int = 240,
    ) -> CheckinRule:
        with SessionLocal() as db:
            rule = CheckinRule(
                latitude=latitude,
                longitude=longitude,
                radius_m=radius_m,
                start_time=start_time,
                grace_minutes=grace_minutes,
                end_time=end_time,
                checkout_grace_minutes=checkout_grace_minutes,
                cross_day_cutoff_minutes=cross_day_cutoff_minutes,
                active=True,
            )
            db.add(rule)
            db.commit()
            db.refresh(rule)
            return rule


    def _login_tokens(self, email: str, password: str, remember_me: bool = True) -> dict:
        response = self.client.post(
            "/auth/login",
            json={"email": email, "password": password, "remember_me": remember_me},
        )
        self.assertEqual(response.status_code, 200, response.text)
        body = response.json()
        self.assertIn("access_token", body)
        self.assertIn("refresh_token", body)
        return body

    def _login(self, email: str, password: str) -> str:
        return self._login_tokens(email, password)["access_token"]

    def _system_headers(self) -> dict[str, str]:
        return {"X-Exception-Workflow-Key": "test-exception-system-key"}

    def _create_attendance_log(
        self,
        *,
        employee_id: int,
        work_date_value: date,
        happened_at: datetime,
        log_type: str = "IN",
    ) -> AttendanceLog:
        with SessionLocal() as db:
            log = AttendanceLog(
                employee_id=employee_id,
                type=log_type,
                time=happened_at,
                work_date=work_date_value,
                lat=10.7769,
                lng=106.7009,
                is_out_of_range=False,
                punctuality_status="ON_TIME" if log_type == "IN" else None,
                checkout_status="ON_TIME" if log_type == "OUT" else None,
            )
            db.add(log)
            db.commit()
            db.refresh(log)
            return log

    def _request_forgot_password(
        self,
        email: str,
        raw_token: str = "known-reset-token",
    ):
        spy_mail = _SpyMailSender()
        with patch("app.api.auth.get_mail_sender", return_value=spy_mail), patch(
            "app.services.auth.password_reset_service.token_urlsafe",
            return_value=raw_token,
        ):
            response = self.client.post("/auth/forgot-password", json={"email": email})
        return response, spy_mail


    def test_register_login_flow(self) -> None:
        email = "flow_user@example.com"
        password = "user123"

        register_res = self.client.post("/auth/register", json={"email": email, "password": password})
        self.assertEqual(register_res.status_code, 201, register_res.text)
        self.assertEqual(register_res.json()["role"], "USER")

        token = self._login(email, password)

        me_res = self.client.get("/auth/me", headers={"Authorization": f"Bearer {token}"})
        self.assertEqual(me_res.status_code, 200, me_res.text)
        self.assertEqual(me_res.json()["email"], email)

    def test_refresh_token_flow(self) -> None:
        email = "refresh_user@example.com"
        password = "user123"
        self._create_user(email=email, password=password, role="USER")

        login_body = self._login_tokens(email, password, remember_me=True)
        old_refresh = login_body["refresh_token"]
        old_access = login_body["access_token"]

        refresh_res = self.client.post("/auth/refresh", json={"refresh_token": old_refresh})
        self.assertEqual(refresh_res.status_code, 200, refresh_res.text)
        refreshed = refresh_res.json()
        self.assertIn("access_token", refreshed)
        self.assertIn("refresh_token", refreshed)
        self.assertNotEqual(refreshed["access_token"], old_access)
        self.assertNotEqual(refreshed["refresh_token"], old_refresh)

        old_refresh_res = self.client.post("/auth/refresh", json={"refresh_token": old_refresh})
        self.assertEqual(old_refresh_res.status_code, 401, old_refresh_res.text)

    def test_logout_and_logout_all_flow(self) -> None:
        email = "logout_user@example.com"
        password = "user123"
        self._create_user(email=email, password=password, role="USER")

        first_login = self._login_tokens(email, password, remember_me=True)
        second_login = self._login_tokens(email, password, remember_me=True)

        logout_res = self.client.post("/auth/logout", json={"refresh_token": first_login["refresh_token"]})
        self.assertEqual(logout_res.status_code, 200, logout_res.text)

        refresh_after_logout = self.client.post(
            "/auth/refresh",
            json={"refresh_token": first_login["refresh_token"]},
        )
        self.assertEqual(refresh_after_logout.status_code, 401, refresh_after_logout.text)

        logout_all_res = self.client.post(
            "/auth/logout-all",
            headers={"Authorization": f"Bearer {second_login['access_token']}"},
        )
        self.assertEqual(logout_all_res.status_code, 200, logout_all_res.text)

        refresh_after_logout_all = self.client.post(
            "/auth/refresh",
            json={"refresh_token": second_login["refresh_token"]},
        )
        self.assertEqual(refresh_after_logout_all.status_code, 401, refresh_after_logout_all.text)

        with SessionLocal() as db:
            user = db.query(User).filter(User.email == email).first()
            self.assertIsNotNone(user)
            active_count = (
                db.query(RefreshToken)
                .filter(
                    RefreshToken.user_id == user.id,
                    RefreshToken.revoked_at.is_(None),
                )
                .count()
            )
            self.assertEqual(active_count, 0)

    def test_forgot_reset_password_happy_path(self) -> None:
        email = "forgot_user@example.com"
        old_password = "oldpass123"
        new_password = "newpass456"
        self._create_user(email=email, password=old_password, role="USER")

        # Create a refresh token before reset to verify revoke-all behavior.
        self._login_tokens(email, old_password, remember_me=True)

        forgot_res, spy_mail = self._request_forgot_password(email=email, raw_token="happy-reset-token")
        self.assertEqual(forgot_res.status_code, 200, forgot_res.text)
        self.assertEqual(
            forgot_res.json().get("message"),
            "Nếu email tồn tại, hệ thống đã gửi hướng dẫn đặt lại mật khẩu.",
        )
        self.assertEqual(len(spy_mail.sent), 1)
        self.assertIn("token=happy-reset-token", spy_mail.sent[0].reset_url)
        self.assertEqual(spy_mail.sent[0].reset_token, "happy-reset-token")

        with SessionLocal() as db:
            user = db.query(User).filter(User.email == email).first()
            self.assertIsNotNone(user)
            token_row = db.query(PasswordResetToken).filter(PasswordResetToken.user_id == user.id).first()
            self.assertIsNotNone(token_row)
            self.assertEqual(token_row.token_hash, hash_token("happy-reset-token"))
            self.assertIsNone(token_row.used_at)

        reset_res = self.client.post(
            "/auth/reset-password",
            json={"token": "happy-reset-token", "new_password": new_password},
        )
        self.assertEqual(reset_res.status_code, 200, reset_res.text)

        old_login = self.client.post(
            "/auth/login",
            json={"email": email, "password": old_password, "remember_me": True},
        )
        self.assertEqual(old_login.status_code, 401, old_login.text)

        new_login = self.client.post(
            "/auth/login",
            json={"email": email, "password": new_password, "remember_me": True},
        )
        self.assertEqual(new_login.status_code, 200, new_login.text)

        with SessionLocal() as db:
            user = db.query(User).filter(User.email == email).first()
            self.assertIsNotNone(user)
            token_row = db.query(PasswordResetToken).filter(PasswordResetToken.user_id == user.id).first()
            self.assertIsNotNone(token_row)
            self.assertIsNotNone(token_row.used_at)
            active_refresh_count = (
                db.query(RefreshToken)
                .filter(RefreshToken.user_id == user.id, RefreshToken.revoked_at.is_(None))
                .count()
            )
            self.assertEqual(active_refresh_count, 1)

    def test_reset_password_rejects_expired_token(self) -> None:
        email = "expired_reset@example.com"
        self._create_user(email=email, password="oldpass123", role="USER")

        self._request_forgot_password(email=email, raw_token="expired-reset-token")

        with SessionLocal() as db:
            user = db.query(User).filter(User.email == email).first()
            self.assertIsNotNone(user)
            token_row = db.query(PasswordResetToken).filter(PasswordResetToken.user_id == user.id).first()
            self.assertIsNotNone(token_row)
            token_row.expires_at = datetime.now(timezone.utc) - timedelta(minutes=1)
            db.commit()

        reset_res = self.client.post(
            "/auth/reset-password",
            json={"token": "expired-reset-token", "new_password": "newpass456"},
        )
        self.assertEqual(reset_res.status_code, 400, reset_res.text)
        self.assertEqual(reset_res.json()["error"]["code"], "INVALID_RESET_TOKEN")

    def test_reset_password_rejects_used_token(self) -> None:
        email = "used_reset@example.com"
        self._create_user(email=email, password="oldpass123", role="USER")

        self._request_forgot_password(email=email, raw_token="used-reset-token")

        with SessionLocal() as db:
            user = db.query(User).filter(User.email == email).first()
            self.assertIsNotNone(user)
            token_row = db.query(PasswordResetToken).filter(PasswordResetToken.user_id == user.id).first()
            self.assertIsNotNone(token_row)
            token_row.used_at = datetime.now(timezone.utc)
            db.commit()

        reset_res = self.client.post(
            "/auth/reset-password",
            json={"token": "used-reset-token", "new_password": "newpass456"},
        )
        self.assertEqual(reset_res.status_code, 400, reset_res.text)
        self.assertEqual(reset_res.json()["error"]["code"], "INVALID_RESET_TOKEN")

    def test_reset_password_rejects_empty_token(self) -> None:
        reset_res = self.client.post(
            "/auth/reset-password",
            json={"token": "   ", "new_password": "newpass456"},
        )
        self.assertEqual(reset_res.status_code, 400, reset_res.text)
        self.assertEqual(reset_res.json()["error"]["code"], "INVALID_RESET_TOKEN")

    def test_reset_password_rejects_malformed_link_token(self) -> None:
        reset_res = self.client.post(
            "/auth/reset-password",
            json={
                "token": "http://localhost:62601/#/reset-password?token=",
                "new_password": "newpass456",
            },
        )
        self.assertEqual(reset_res.status_code, 400, reset_res.text)
        self.assertEqual(reset_res.json()["error"]["code"], "INVALID_RESET_TOKEN")

    def test_password_reset_url_builder_handles_malformed_base(self) -> None:
        with patch(
            "app.services.auth.password_reset_service.settings.RESET_PASSWORD_URL_BASE",
            "http://localhost:62601/?foo=1#/reset-password?token=",
        ):
            reset_url = PasswordResetService._build_reset_url("abc-token")

        self.assertIn("#/reset-password?token=abc-token", reset_url)
        self.assertNotIn("?token=#/reset-password", reset_url)
        self.assertNotIn("token=&", reset_url)

    def test_password_reset_url_builder_path_base_keeps_token_for_hash_router(self) -> None:
        with patch(
            "app.services.auth.password_reset_service.settings.RESET_PASSWORD_URL_BASE",
            "https://chamcongweb-uat.vercel.app/reset-password",
        ):
            reset_url = PasswordResetService._build_reset_url("abc-token")

        self.assertIn("/reset-password?token=abc-token", reset_url)
        self.assertIn("#/reset-password?token=abc-token", reset_url)

    def test_password_reset_cleanup_removes_expired_and_old_used_tokens(self) -> None:
        user = self._create_user(email="cleanup_reset@example.com", password="oldpass123", role="USER")
        now_utc = datetime(2026, 3, 19, 0, 0, tzinfo=timezone.utc)

        with SessionLocal() as db:
            db.add_all(
                [
                    PasswordResetToken(
                        user_id=user.id,
                        token_hash=hash_token("expired-token"),
                        expires_at=now_utc - timedelta(minutes=5),
                        used_at=None,
                    ),
                    PasswordResetToken(
                        user_id=user.id,
                        token_hash=hash_token("used-old-token"),
                        expires_at=now_utc + timedelta(days=1),
                        used_at=now_utc - timedelta(days=2),
                    ),
                    PasswordResetToken(
                        user_id=user.id,
                        token_hash=hash_token("active-token"),
                        expires_at=now_utc + timedelta(days=1),
                        used_at=None,
                    ),
                ]
            )
            db.commit()

            deleted = cleanup_password_reset_tokens(
                db,
                now_utc=now_utc,
                used_retention_days=1,
            )
            db.commit()
            self.assertEqual(deleted, 2)

            remaining_hashes = {row.token_hash for row in db.query(PasswordResetToken).all()}
            self.assertEqual(remaining_hashes, {hash_token("active-token")})


    def test_set_rule_flow(self) -> None:
        self._create_user(email="admin@example.com", password="admin123", role="ADMIN")
        admin_token = self._login("admin@example.com", "admin123")

        put_res = self.client.put(
            "/rules/active",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={
                "lat": 10.7769,
                "lng": 106.7009,
                "radius": 250,
                "start_time": "08:00",
                "grace_minutes": 30,
                "end_time": "17:30",
                "checkout_grace_minutes": 10,
            },
        )
        self.assertEqual(put_res.status_code, 200, put_res.text)
        self.assertEqual(put_res.json()["radius_m"], 250)
        self.assertEqual(put_res.json()["end_time"], "17:30")
        self.assertEqual(put_res.json()["checkout_grace_minutes"], 10)
        self.assertIsNone(put_res.json().get("radius_policy_warning"))

        warn_res = self.client.put(
            "/rules/active",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={
                "lat": 10.7769,
                "lng": 106.7009,
                "radius": 350,
                "start_time": "08:00",
                "grace_minutes": 30,
                "end_time": "17:30",
                "checkout_grace_minutes": 10,
            },
        )
        self.assertEqual(warn_res.status_code, 200, warn_res.text)
        self.assertEqual(warn_res.json()["radius_m"], 350)
        self.assertEqual(
            warn_res.json().get("radius_policy_warning"),
            "RADIUS_ABOVE_POLICY_THRESHOLD",
        )

        invalid_res = self.client.put(
            "/rules/active",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={
                "lat": 10.7769,
                "lng": 106.7009,
                "radius": 10,
                "start_time": "08:00",
                "grace_minutes": 30,
                "end_time": "17:30",
                "checkout_grace_minutes": 10,
            },
        )
        self.assertEqual(invalid_res.status_code, 422, invalid_res.text)

        get_res = self.client.get(
            "/rules/active",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        self.assertEqual(get_res.status_code, 200, get_res.text)
        self.assertEqual(get_res.json()["radius_m"], 350)
        self.assertEqual(
            get_res.json().get("radius_policy_warning"),
            "RADIUS_ABOVE_POLICY_THRESHOLD",
        )

    def test_group_time_rule_crud_flow(self) -> None:
        self._create_user(email="admin_group_time@example.com", password="admin123", role="ADMIN")
        admin_token = self._login("admin_group_time@example.com", "admin123")
        headers = {"Authorization": f"Bearer {admin_token}"}

        create_res = self.client.post(
            "/groups",
            headers=headers,
            json={
                "code": "GTM_API",
                "name": "Group Time API",
                "start_time": "09:00",
                "grace_minutes": 10,
                "end_time": "18:00",
                "checkout_grace_minutes": 5,
                "active": True,
            },
        )
        self.assertEqual(create_res.status_code, 200, create_res.text)
        body = create_res.json()
        self.assertEqual(body["start_time"], "09:00")
        self.assertEqual(body["grace_minutes"], 10)
        self.assertEqual(body["end_time"], "18:00")
        self.assertEqual(body["checkout_grace_minutes"], 5)

        group_id = body["id"]
        update_res = self.client.put(
            f"/groups/{group_id}",
            headers=headers,
            json={
                "grace_minutes": 25,
                "checkout_grace_minutes": 20,
                "start_time": None,
            },
        )
        self.assertEqual(update_res.status_code, 200, update_res.text)
        updated = update_res.json()
        self.assertIsNone(updated["start_time"])
        self.assertEqual(updated["grace_minutes"], 25)
        self.assertEqual(updated["checkout_grace_minutes"], 20)

    def test_group_geofence_radius_guardrail_and_warning(self) -> None:
        self._create_user(email="admin_geofence_policy@example.com", password="admin123", role="ADMIN")
        admin_token = self._login("admin_geofence_policy@example.com", "admin123")
        headers = {"Authorization": f"Bearer {admin_token}"}

        create_group_res = self.client.post(
            "/groups",
            headers=headers,
            json={
                "code": "G_RADIUS",
                "name": "Group Radius",
                "active": True,
            },
        )
        self.assertEqual(create_group_res.status_code, 200, create_group_res.text)
        group_id = create_group_res.json()["id"]

        too_small_res = self.client.post(
            f"/groups/{group_id}/geofences",
            headers=headers,
            json={
                "name": "Too Small",
                "latitude": 10.7769,
                "longitude": 106.7009,
                "radius_m": 10,
                "active": True,
            },
        )
        self.assertEqual(too_small_res.status_code, 422, too_small_res.text)

        warning_res = self.client.post(
            f"/groups/{group_id}/geofences",
            headers=headers,
            json={
                "name": "Warning Radius",
                "latitude": 10.7769,
                "longitude": 106.7009,
                "radius_m": 350,
                "active": True,
            },
        )
        self.assertEqual(warning_res.status_code, 200, warning_res.text)
        self.assertEqual(warning_res.json()["radius_m"], 350)
        self.assertEqual(
            warning_res.json().get("radius_policy_warning"),
            "RADIUS_ABOVE_POLICY_THRESHOLD",
        )

        list_res = self.client.get(
            f"/groups/{group_id}/geofences",
            headers=headers,
        )
        self.assertEqual(list_res.status_code, 200, list_res.text)
        self.assertEqual(list_res.json()[0]["radius_m"], 350)
        self.assertEqual(
            list_res.json()[0].get("radius_policy_warning"),
            "RADIUS_ABOVE_POLICY_THRESHOLD",
        )

    def test_checkin_checkout_flow(self) -> None:
        user = self._create_user(email="staff@example.com", password="staff123", role="USER")
        self._create_employee(code="EM001", full_name="Staff One", user_id=user.id)
        self._create_rule()

        token = self._login("staff@example.com", "staff123")
        headers = {"Authorization": f"Bearer {token}"}

        checkin_res = self.client.post(
            "/attendance/checkin",
            headers=headers,
            json={"lat": 10.7769, "lng": 106.7009},
        )
        self.assertEqual(checkin_res.status_code, 200, checkin_res.text)
        self.assertEqual(checkin_res.json()["log"]["type"], "IN")
        self.assertIn(checkin_res.json()["log"].get("punctuality_status"), {"EARLY", "ON_TIME", "LATE"})

        checkout_res = self.client.post(
            "/attendance/checkout",
            headers=headers,
            json={"lat": 10.7769, "lng": 106.7009},
        )
        self.assertEqual(checkout_res.status_code, 200, checkout_res.text)
        self.assertEqual(checkout_res.json()["log"]["type"], "OUT")
        self.assertIn(checkout_res.json()["log"].get("checkout_status"), {"EARLY", "ON_TIME", "LATE"})

        repeat_checkin_same_day = self.client.post(
            "/attendance/checkin",
            headers=headers,
            json={"lat": 10.7769, "lng": 106.7009},
        )
        self.assertEqual(repeat_checkin_same_day.status_code, 400, repeat_checkin_same_day.text)

        status_res = self.client.get("/attendance/status", headers=headers)
        self.assertEqual(status_res.status_code, 200, status_res.text)
        status_body = status_res.json()
        self.assertFalse(status_body["can_checkin"])
        self.assertFalse(status_body["can_checkout"])


    def test_group_geofence_flow(self) -> None:
        user = self._create_user(email="group_user@example.com", password="user123", role="USER")

        group_a = self._create_group("GA", "Group A")
        group_b = self._create_group("GB", "Group B")

        self._create_geofence(group_a.id, "Gate", 10.7769, 106.7009, 200)
        self._create_geofence(group_a.id, "Annex", 10.7774, 106.7014, 200)
        self._create_geofence(group_b.id, "Warehouse", 10.7905, 106.5950, 300)

        self._create_employee(code="EM003", full_name="Group User", user_id=user.id, group_id=group_a.id)
        self._create_rule()  # fallback + timing source

        token = self._login("group_user@example.com", "user123")
        headers = {"Authorization": f"Bearer {token}"}

        checkin_res = self.client.post(
            "/attendance/checkin",
            headers=headers,
            json={"lat": 10.7774, "lng": 106.7014},
        )
        self.assertEqual(checkin_res.status_code, 200, checkin_res.text)
        self.assertFalse(checkin_res.json()["log"]["is_out_of_range"])
        self.assertEqual(checkin_res.json()["geofence_source"], "GROUP")
        self.assertIsNone(checkin_res.json()["fallback_reason"])

        checkout_res = self.client.post(
            "/attendance/checkout",
            headers=headers,
            json={"lat": 10.7905, "lng": 106.5950},
        )
        self.assertEqual(checkout_res.status_code, 200, checkout_res.text)
        self.assertTrue(checkout_res.json()["log"]["is_out_of_range"])
        self.assertEqual(checkout_res.json()["geofence_source"], "GROUP")
        self.assertIsNone(checkout_res.json()["fallback_reason"])

    def test_out_of_range_gps_risk_creates_suspected_spoof_exception_once(self) -> None:
        admin = self._create_user(email="admin_gps_risk@example.com", password="admin123", role="ADMIN")
        user = self._create_user(email="gps_risk_user@example.com", password="user123", role="USER")

        group = self._create_group("GPSR", "GPS Risk Group")
        self._create_geofence(group.id, "Office Gate", 10.7769, 106.7009, 200)
        employee = self._create_employee(code="EMGPS", full_name="GPS Risk User", user_id=user.id, group_id=group.id)
        self._create_rule(latitude=10.7769, longitude=106.7009, radius_m=300)

        user_headers = {"Authorization": f"Bearer {create_access_token({'sub': str(user.id), 'role': user.role})}"}
        admin_headers = {"Authorization": f"Bearer {create_access_token({'sub': str(admin.id), 'role': admin.role})}"}

        _FixedDateTime.fixed_now = datetime(2026, 3, 12, 1, 0, tzinfo=timezone.utc)
        with patch("app.api.attendance.datetime", _FixedDateTime):
            checkin_res = self.client.post(
                "/attendance/checkin",
                headers=user_headers,
                json={"lat": 10.7905, "lng": 106.5950},
            )
        self.assertEqual(checkin_res.status_code, 200, checkin_res.text)
        self.assertTrue(checkin_res.json()["log"]["is_out_of_range"])
        self.assertEqual(checkin_res.json()["decision"], "ALLOW_WITH_EXCEPTION")

        employee_exceptions_res = self.client.get(
            "/reports/attendance-exceptions/me?status=PENDING_EMPLOYEE",
            headers=user_headers,
        )
        self.assertEqual(employee_exceptions_res.status_code, 200, employee_exceptions_res.text)
        employee_rows = [
            row for row in employee_exceptions_res.json()
            if row["employee_code"] == "EMGPS"
        ]
        self.assertEqual(len(employee_rows), 1)
        self.assertEqual(employee_rows[0]["exception_type"], "SUSPECTED_LOCATION_SPOOF")
        self.assertEqual(employee_rows[0]["status"], "PENDING_EMPLOYEE")

        admin_exceptions_res = self.client.get(
            "/reports/attendance-exceptions?from=2026-03-12&to=2026-03-12&exception_type=SUSPECTED_LOCATION_SPOOF&status=PENDING_EMPLOYEE",
            headers=admin_headers,
        )
        self.assertEqual(admin_exceptions_res.status_code, 200, admin_exceptions_res.text)
        admin_rows = [
            row for row in admin_exceptions_res.json()
            if row["employee_code"] == "EMGPS"
        ]
        self.assertEqual(len(admin_rows), 1)

        _FixedDateTime.fixed_now = datetime(2026, 3, 12, 10, 0, tzinfo=timezone.utc)
        with patch("app.api.attendance.datetime", _FixedDateTime):
            checkout_res = self.client.post(
                "/attendance/checkout",
                headers=user_headers,
                json={"lat": 10.7905, "lng": 106.5950},
            )
        self.assertEqual(checkout_res.status_code, 200, checkout_res.text)
        self.assertEqual(checkout_res.json()["decision"], "ALLOW_WITH_EXCEPTION")

        with SessionLocal() as db:
            exceptions = (
                db.query(AttendanceException)
                .filter(AttendanceException.employee_id == employee.id)
                .all()
            )
        self.assertEqual(len(exceptions), 1)
        self.assertEqual(exceptions[0].exception_type, "SUSPECTED_LOCATION_SPOOF")
        self.assertEqual(exceptions[0].status, "PENDING_EMPLOYEE")

        _FixedDateTime.fixed_now = None

    def test_group_time_rule_overrides_system_rule(self) -> None:
        user = self._create_user(email="group_time_user@example.com", password="user123", role="USER")
        group = self._create_group(
            "GTIME",
            "Group Time",
            start_time=time(9, 0),
            grace_minutes=5,
            end_time=time(17, 0),
            checkout_grace_minutes=0,
        )
        self._create_geofence(group.id, "Time Gate", 10.7769, 106.7009, 250)
        self._create_employee(code="EM006", full_name="Group Time User", user_id=user.id, group_id=group.id)

        # System fallback rule remains different, so we can assert override behavior.
        self._create_rule(latitude=10.7769, longitude=106.7009, radius_m=300)

        token = self._login("group_time_user@example.com", "user123")
        headers = {"Authorization": f"Bearer {token}"}

        _FixedDateTime.fixed_now = datetime(2026, 3, 11, 1, 5, tzinfo=timezone.utc)  # 08:05 VN
        with patch("app.api.attendance.datetime", _FixedDateTime):
            checkin_res = self.client.post(
                "/attendance/checkin",
                headers=headers,
                json={"lat": 10.7769, "lng": 106.7009},
            )
        self.assertEqual(checkin_res.status_code, 200, checkin_res.text)
        self.assertEqual(checkin_res.json()["log"]["punctuality_status"], "EARLY")

        _FixedDateTime.fixed_now = datetime(2026, 3, 11, 10, 10, tzinfo=timezone.utc)  # 17:10 VN
        with patch("app.api.attendance.datetime", _FixedDateTime):
            checkout_res = self.client.post(
                "/attendance/checkout",
                headers=headers,
                json={"lat": 10.7769, "lng": 106.7009},
            )
        self.assertEqual(checkout_res.status_code, 200, checkout_res.text)
        self.assertEqual(checkout_res.json()["log"]["checkout_status"], "LATE")

        _FixedDateTime.fixed_now = None

    def test_group_cutoff_overrides_system_cutoff(self) -> None:
        user = self._create_user(email="group_cutoff_user@example.com", password="user123", role="USER")
        group = self._create_group(
            "GCUT",
            "Group Cutoff",
            start_time=time(8, 0),
            grace_minutes=30,
            end_time=time(17, 0),
            checkout_grace_minutes=0,
            cross_day_cutoff_minutes=360,
        )
        self._create_geofence(group.id, "Cutoff Gate", 10.7769, 106.7009, 250)
        self._create_employee(code="EM013", full_name="Group Cutoff User", user_id=user.id, group_id=group.id)

        self._create_rule(latitude=10.7769, longitude=106.7009, radius_m=300, cross_day_cutoff_minutes=240)

        token = self._login("group_cutoff_user@example.com", "user123")
        headers = {"Authorization": f"Bearer {token}"}

        _FixedDateTime.fixed_now = datetime(2026, 3, 10, 17, 30, tzinfo=timezone.utc)  # 00:30 VN
        with patch("app.api.attendance.datetime", _FixedDateTime):
            in_res = self.client.post(
                "/attendance/checkin",
                headers=headers,
                json={"lat": 10.7769, "lng": 106.7009},
            )
        self.assertEqual(in_res.status_code, 200, in_res.text)

        _FixedDateTime.fixed_now = datetime(2026, 3, 10, 22, 0, tzinfo=timezone.utc)  # 05:00 VN
        with patch("app.api.attendance.datetime", _FixedDateTime):
            status_before_cutoff = self.client.get("/attendance/status", headers=headers)
        self.assertEqual(status_before_cutoff.status_code, 200, status_before_cutoff.text)
        self.assertTrue(status_before_cutoff.json()["can_checkout"])
        self.assertFalse(status_before_cutoff.json()["can_checkin"])

        _FixedDateTime.fixed_now = datetime(2026, 3, 10, 23, 30, tzinfo=timezone.utc)  # 06:30 VN > group cutoff
        with patch("app.api.attendance.datetime", _FixedDateTime):
            status_after_cutoff = self.client.get("/attendance/status", headers=headers)
        self.assertEqual(status_after_cutoff.status_code, 200, status_after_cutoff.text)
        self.assertEqual(status_after_cutoff.json()["warning_code"], "AUTO_CLOSED")
        self.assertTrue(status_after_cutoff.json()["can_checkin"])

        _FixedDateTime.fixed_now = None

    def test_employee_without_group_uses_active_rule_fallback(self) -> None:
        user = self._create_user(email="nogroup_user@example.com", password="user123", role="USER")
        self._create_employee(code="EM004", full_name="No Group User", user_id=user.id, group_id=None)
        self._create_rule(latitude=10.7769, longitude=106.7009, radius_m=300)

        token = self._login("nogroup_user@example.com", "user123")
        headers = {"Authorization": f"Bearer {token}"}

        checkin_res = self.client.post(
            "/attendance/checkin",
            headers=headers,
            json={"lat": 10.7769, "lng": 106.7009},
        )
        self.assertEqual(checkin_res.status_code, 200, checkin_res.text)
        self.assertFalse(checkin_res.json()["log"]["is_out_of_range"])
        self.assertEqual(checkin_res.json()["log"]["matched_geofence"], "SYSTEM_RULE")
        self.assertEqual(checkin_res.json()["geofence_source"], "SYSTEM_FALLBACK")
        self.assertEqual(checkin_res.json()["fallback_reason"], "EMPLOYEE_NOT_ASSIGNED_GROUP")

        checkout_res = self.client.post(
            "/attendance/checkout",
            headers=headers,
            json={"lat": 10.7905, "lng": 106.5950},
        )
        self.assertEqual(checkout_res.status_code, 200, checkout_res.text)
        self.assertTrue(checkout_res.json()["log"]["is_out_of_range"])
        self.assertIsNone(checkout_res.json()["log"]["matched_geofence"])
        self.assertEqual(checkout_res.json()["geofence_source"], "SYSTEM_FALLBACK")
        self.assertEqual(checkout_res.json()["fallback_reason"], "EMPLOYEE_NOT_ASSIGNED_GROUP")

    def test_daily_report_contains_fallback_source(self) -> None:
        admin = self._create_user(email="admin_daily_report@example.com", password="admin123", role="ADMIN")
        user = self._create_user(email="daily_report_user@example.com", password="user123", role="USER")

        self._create_employee(code="EM009", full_name="Daily Report User", user_id=user.id, group_id=None)
        self._create_rule(latitude=10.7769, longitude=106.7009, radius_m=300)

        user_token = self._login("daily_report_user@example.com", "user123")
        user_headers = {"Authorization": f"Bearer {user_token}"}

        in_res = self.client.post(
            "/attendance/checkin",
            headers=user_headers,
            json={"lat": 10.7769, "lng": 106.7009},
        )
        self.assertEqual(in_res.status_code, 200, in_res.text)

        out_res = self.client.post(
            "/attendance/checkout",
            headers=user_headers,
            json={"lat": 10.7905, "lng": 106.5950},
        )
        self.assertEqual(out_res.status_code, 200, out_res.text)

        admin_token = self._login("admin_daily_report@example.com", "admin123")
        report_res = self.client.get(
            "/attendance/report/daily",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        self.assertEqual(report_res.status_code, 200, report_res.text)

        rows = report_res.json()
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["employee_code"], "EM009")
        self.assertEqual(rows[0]["geofence_source"], "SYSTEM_FALLBACK")
        self.assertEqual(rows[0]["fallback_reason"], "EMPLOYEE_NOT_ASSIGNED_GROUP")
    def test_group_inactive_uses_system_fallback(self) -> None:
        user = self._create_user(email="inactive_group_user@example.com", password="user123", role="USER")
        group = self._create_group("GINACT", "Inactive Group", active=False)
        self._create_geofence(group.id, "Inactive Gate", 10.7769, 106.7009, 500, active=True)
        self._create_employee(code="EM007", full_name="Inactive Group User", user_id=user.id, group_id=group.id)
        self._create_rule(latitude=10.7769, longitude=106.7009, radius_m=300)

        token = self._login("inactive_group_user@example.com", "user123")
        headers = {"Authorization": f"Bearer {token}"}

        checkin_res = self.client.post(
            "/attendance/checkin",
            headers=headers,
            json={"lat": 10.7769, "lng": 106.7009},
        )
        self.assertEqual(checkin_res.status_code, 200, checkin_res.text)
        self.assertEqual(checkin_res.json()["geofence_source"], "SYSTEM_FALLBACK")
        self.assertEqual(checkin_res.json()["fallback_reason"], "GROUP_INACTIVE_OR_NOT_FOUND")

    def test_group_without_active_geofence_uses_system_fallback(self) -> None:
        user = self._create_user(email="no_active_geofence_user@example.com", password="user123", role="USER")
        group = self._create_group("GNOFG", "No Active Geofence Group", active=True)
        self._create_geofence(group.id, "Inactive Geofence", 10.7769, 106.7009, 500, active=False)
        self._create_employee(code="EM008", full_name="No Active Geofence User", user_id=user.id, group_id=group.id)
        self._create_rule(latitude=10.7769, longitude=106.7009, radius_m=300)

        token = self._login("no_active_geofence_user@example.com", "user123")
        headers = {"Authorization": f"Bearer {token}"}

        checkin_res = self.client.post(
            "/attendance/checkin",
            headers=headers,
            json={"lat": 10.7769, "lng": 106.7009},
        )
        self.assertEqual(checkin_res.status_code, 200, checkin_res.text)
        self.assertEqual(checkin_res.json()["geofence_source"], "SYSTEM_FALLBACK")
        self.assertEqual(checkin_res.json()["fallback_reason"], "NO_ACTIVE_GEOFENCE_IN_GROUP")

    def test_delete_group_flow(self) -> None:
        admin = self._create_user(email="admin_group@example.com", password="admin123", role="ADMIN")
        user = self._create_user(email="user_group@example.com", password="user123", role="USER")

        group = self._create_group("DEL", "Delete Group")
        self._create_geofence(group.id, "Delete Gate", 10.7769, 106.7009, 300)
        employee = self._create_employee(code="EM005", full_name="Delete User", user_id=user.id, group_id=group.id)

        admin_token = self._login("admin_group@example.com", "admin123")
        res = self.client.delete(
            f"/groups/{group.id}",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        self.assertEqual(res.status_code, 200, res.text)

        with SessionLocal() as db:
            refreshed_emp = db.query(Employee).filter(Employee.id == employee.id).first()
            self.assertIsNotNone(refreshed_emp)
            self.assertIsNone(refreshed_emp.group_id)

            deleted_group = db.query(Group).filter(Group.id == group.id).first()
            self.assertIsNone(deleted_group)

            deleted_geofence = db.query(GroupGeofence).filter(GroupGeofence.group_id == group.id).first()
            self.assertIsNone(deleted_geofence)

    def test_continuous_session_ot_cross_day_split_minutes(self) -> None:
        admin = self._create_user(email="admin_ot@example.com", password="admin123", role="ADMIN")
        user = self._create_user(email="ot_user@example.com", password="user123", role="USER")

        self._create_employee(code="EM010", full_name="OT User", user_id=user.id)
        self._create_rule(
            latitude=10.7769,
            longitude=106.7009,
            radius_m=300,
            start_time=time(8, 0),
            grace_minutes=30,
            end_time=time(17, 0),
            checkout_grace_minutes=0,
        )

        token = self._login("ot_user@example.com", "user123")
        headers = {"Authorization": f"Bearer {token}"}

        _FixedDateTime.fixed_now = datetime(2026, 3, 10, 1, 0, tzinfo=timezone.utc)  # 08:00 VN
        with patch("app.api.attendance.datetime", _FixedDateTime):
            in_res = self.client.post(
                "/attendance/checkin",
                headers=headers,
                json={"lat": 10.7769, "lng": 106.7009},
            )
        self.assertEqual(in_res.status_code, 200, in_res.text)

        _FixedDateTime.fixed_now = datetime(2026, 3, 10, 19, 30, tzinfo=timezone.utc)  # 02:30 VN (next day)
        with patch("app.api.attendance.datetime", _FixedDateTime):
            out_res = self.client.post(
                "/attendance/checkout",
                headers=headers,
                json={"lat": 10.7769, "lng": 106.7009},
            )
        self.assertEqual(out_res.status_code, 200, out_res.text)

        admin_token = self._login("admin_ot@example.com", "admin123")
        report_res = self.client.get(
            "/attendance/report/daily?from_date=2026-03-10&to_date=2026-03-10",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        self.assertEqual(report_res.status_code, 200, report_res.text)

        rows = report_res.json()
        self.assertTrue(any(row["employee_code"] == "EM010" for row in rows))
        row = next(r for r in rows if r["employee_code"] == "EM010")

        self.assertEqual(row["date"], "2026-03-10")
        self.assertEqual(row["regular_minutes"], 540)
        self.assertEqual(row["overtime_minutes"], 570)
        self.assertEqual(row["payable_overtime_minutes"], 570)
        self.assertTrue(row["overtime_cross_day"])

        _FixedDateTime.fixed_now = None

    def test_auto_close_after_cutoff_creates_exception(self) -> None:
        admin = self._create_user(email="admin_auto_close@example.com", password="admin123", role="ADMIN")
        user = self._create_user(email="auto_close_user@example.com", password="user123", role="USER")

        self._create_employee(code="EM011", full_name="Auto Close User", user_id=user.id)
        self._create_rule(latitude=10.7769, longitude=106.7009, radius_m=300)

        token = self._login("auto_close_user@example.com", "user123")
        headers = {"Authorization": f"Bearer {token}"}

        _FixedDateTime.fixed_now = datetime(2026, 3, 10, 2, 0, tzinfo=timezone.utc)  # 09:00 VN
        with patch("app.api.attendance.datetime", _FixedDateTime):
            in_res = self.client.post(
                "/attendance/checkin",
                headers=headers,
                json={"lat": 10.7769, "lng": 106.7009},
            )
        self.assertEqual(in_res.status_code, 200, in_res.text)

        _FixedDateTime.fixed_now = datetime(2026, 3, 10, 23, 0, tzinfo=timezone.utc)  # 06:00 VN next day > cutoff 04:00
        with patch("app.api.attendance.datetime", _FixedDateTime):
            status_res = self.client.get("/attendance/status", headers=headers)
        self.assertEqual(status_res.status_code, 200, status_res.text)
        self.assertTrue(status_res.json()["can_checkin"])
        self.assertEqual(status_res.json()["warning_code"], "AUTO_CLOSED")

        admin_token = self._login("admin_auto_close@example.com", "admin123")
        report_res = self.client.get(
            "/reports/attendance-exceptions?from=2026-03-10&to=2026-03-10&exception_type=AUTO_CLOSED",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        self.assertEqual(report_res.status_code, 200, report_res.text)
        rows = report_res.json()
        self.assertTrue(any(row["employee_code"] == "EM011" for row in rows))

        daily_report_res = self.client.get(
            "/attendance/report/daily?from_date=2026-03-10&to_date=2026-03-10",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        self.assertEqual(daily_report_res.status_code, 200, daily_report_res.text)
        daily_rows = daily_report_res.json()
        row = next(r for r in daily_rows if r["employee_code"] == "EM011")
        self.assertEqual(row["exception_status"], "PENDING_EMPLOYEE")
        self.assertEqual(row["attendance_state"], "PENDING_TIMESHEET")
        self.assertEqual(row["checkout_status"], "SYSTEM_AUTO")
        self.assertGreater(row["overtime_minutes"], 0)
        self.assertEqual(row["payable_overtime_minutes"], 0)

        _FixedDateTime.fixed_now = None

    def test_missed_checkout_auto_exception_approve_and_block_reopen(self) -> None:
        admin = self._create_user(email="admin_missed@example.com", password="admin123", role="ADMIN")
        user = self._create_user(email="missed_user@example.com", password="user123", role="USER")

        self._create_employee(code="EM015", full_name="Missed Checkout User", user_id=user.id)
        self._create_rule(
            latitude=10.7769,
            longitude=106.7009,
            radius_m=300,
            start_time=time(8, 0),
            grace_minutes=30,
            end_time=time(17, 0),
            checkout_grace_minutes=0,
            cross_day_cutoff_minutes=240,
        )

        token = self._login("missed_user@example.com", "user123")
        headers = {"Authorization": f"Bearer {token}"}

        _FixedDateTime.fixed_now = datetime(2026, 3, 10, 1, 0, tzinfo=timezone.utc)  # 08:00 VN
        with patch("app.api.attendance.datetime", _FixedDateTime):
            in_res = self.client.post(
                "/attendance/checkin",
                headers=headers,
                json={"lat": 10.7769, "lng": 106.7009},
            )
        self.assertEqual(in_res.status_code, 200, in_res.text)

        _FixedDateTime.fixed_now = datetime(2026, 3, 10, 11, 30, tzinfo=timezone.utc)  # 18:30 VN
        with patch("app.api.attendance.datetime", _FixedDateTime):
            status_res = self.client.get("/attendance/status", headers=headers)
        self.assertEqual(status_res.status_code, 200, status_res.text)
        self.assertEqual(status_res.json()["warning_code"], "MISSED_CHECKOUT")
        self.assertEqual(status_res.json()["warning_date"], "2026-03-10")
        self.assertTrue(status_res.json()["can_checkout"])

        admin_token = self._login("admin_missed@example.com", "admin123")
        admin_headers = {"Authorization": f"Bearer {admin_token}"}

        exception_res = self.client.get(
            "/reports/attendance-exceptions?from=2026-03-10&to=2026-03-10&exception_type=MISSED_CHECKOUT",
            headers=admin_headers,
        )
        self.assertEqual(exception_res.status_code, 200, exception_res.text)
        rows = [r for r in exception_res.json() if r["employee_code"] == "EM015"]
        self.assertTrue(rows)
        exception_id = rows[0]["id"]

        resolve_missing_time_res = self.client.post(
            f"/reports/attendance-exceptions/{exception_id}/approve",
            headers=admin_headers,
            json={},
        )
        self.assertEqual(resolve_missing_time_res.status_code, 409, resolve_missing_time_res.text)

        submit_explanation_res = self.client.post(
            f"/reports/attendance-exceptions/{exception_id}/submit-explanation",
            headers=headers,
            json={"explanation": "Quen bam checkout sau khi roi cong ty"},
        )
        self.assertEqual(submit_explanation_res.status_code, 200, submit_explanation_res.text)
        submitted = submit_explanation_res.json()
        self.assertEqual(submitted["status"], "PENDING_ADMIN")
        self.assertEqual(
            submitted["employee_explanation"],
            "Quen bam checkout sau khi roi cong ty",
        )

        approve_ok_res = self.client.post(
            f"/reports/attendance-exceptions/{exception_id}/approve",
            headers=admin_headers,
            json={
                "admin_note": "Admin entered actual checkout",
                "actual_checkout_time": "2026-03-10T10:30:00+00:00",
            },
        )
        self.assertEqual(approve_ok_res.status_code, 200, approve_ok_res.text)
        resolved = approve_ok_res.json()
        self.assertEqual(resolved["status"], "APPROVED")
        self.assertEqual(resolved["resolved_by"], admin.id)
        self.assertIsNotNone(resolved["actual_checkout_time"])
        self.assertEqual(resolved["admin_note"], "Admin entered actual checkout")

        daily_resolved = self.client.get(
            "/attendance/report/daily?from_date=2026-03-10&to_date=2026-03-10",
            headers=admin_headers,
        )
        self.assertEqual(daily_resolved.status_code, 200, daily_resolved.text)
        resolved_row = next(r for r in daily_resolved.json() if r["employee_code"] == "EM015")
        self.assertEqual(resolved_row["exception_status"], "APPROVED")
        self.assertEqual(resolved_row["attendance_state"], "COMPLETE")
        self.assertEqual(resolved_row["checkout_status"], "LATE")
        self.assertEqual(resolved_row["regular_minutes"], 540)
        self.assertEqual(resolved_row["overtime_minutes"], 30)
        self.assertEqual(resolved_row["payable_overtime_minutes"], 30)

        reopen_res = self.client.patch(
            f"/reports/attendance-exceptions/{exception_id}/reopen",
            headers=admin_headers,
        )
        self.assertEqual(reopen_res.status_code, 409, reopen_res.text)

        _FixedDateTime.fixed_now = None
    def test_auto_closed_exception_approve_and_block_reopen(self) -> None:
        admin = self._create_user(email="admin_resolve@example.com", password="admin123", role="ADMIN")
        user = self._create_user(email="resolve_user@example.com", password="user123", role="USER")

        self._create_employee(code="EM013", full_name="Resolve User", user_id=user.id)
        self._create_rule(
            latitude=10.7769,
            longitude=106.7009,
            radius_m=300,
            start_time=time(8, 0),
            grace_minutes=30,
            end_time=time(17, 0),
            checkout_grace_minutes=0,
        )

        token = self._login("resolve_user@example.com", "user123")
        headers = {"Authorization": f"Bearer {token}"}

        _FixedDateTime.fixed_now = datetime(2026, 3, 10, 1, 0, tzinfo=timezone.utc)  # 08:00 VN
        with patch("app.api.attendance.datetime", _FixedDateTime):
            in_res = self.client.post(
                "/attendance/checkin",
                headers=headers,
                json={"lat": 10.7769, "lng": 106.7009},
            )
        self.assertEqual(in_res.status_code, 200, in_res.text)

        _FixedDateTime.fixed_now = datetime(2026, 3, 10, 23, 30, tzinfo=timezone.utc)  # 06:30 VN next day > cutoff
        with patch("app.api.attendance.datetime", _FixedDateTime):
            status_res = self.client.get("/attendance/status", headers=headers)
        self.assertEqual(status_res.status_code, 200, status_res.text)
        self.assertEqual(status_res.json()["warning_code"], "AUTO_CLOSED")

        admin_token = self._login("admin_resolve@example.com", "admin123")
        admin_headers = {"Authorization": f"Bearer {admin_token}"}

        exception_res = self.client.get(
            "/reports/attendance-exceptions?from=2026-03-10&to=2026-03-10&exception_type=AUTO_CLOSED",
            headers=admin_headers,
        )
        self.assertEqual(exception_res.status_code, 200, exception_res.text)
        rows = [r for r in exception_res.json() if r["employee_code"] == "EM013"]
        self.assertTrue(rows)
        exception_id = rows[0]["id"]

        resolve_missing_time_res = self.client.post(
            f"/reports/attendance-exceptions/{exception_id}/approve",
            headers=admin_headers,
            json={},
        )
        self.assertEqual(resolve_missing_time_res.status_code, 409, resolve_missing_time_res.text)


        submit_explanation_res = self.client.post(
            f"/reports/attendance-exceptions/{exception_id}/submit-explanation",
            headers=headers,
            json={"explanation": "Da checkout thuc te truoc khi he thong auto close"},
        )
        self.assertEqual(submit_explanation_res.status_code, 200, submit_explanation_res.text)
        self.assertEqual(submit_explanation_res.json()["status"], "PENDING_ADMIN")

        resolve_res = self.client.post(
            f"/reports/attendance-exceptions/{exception_id}/approve",
            headers=admin_headers,
            json={
                "admin_note": "Admin confirmed checkout time",
                "actual_checkout_time": "2026-03-10T10:00:00+00:00",
            },
        )
        self.assertEqual(resolve_res.status_code, 200, resolve_res.text)
        resolved = resolve_res.json()
        self.assertEqual(resolved["status"], "APPROVED")
        self.assertEqual(resolved["resolved_by"], admin.id)
        self.assertIsNotNone(resolved["actual_checkout_time"])

        daily_resolved = self.client.get(
            "/attendance/report/daily?from_date=2026-03-10&to_date=2026-03-10",
            headers=admin_headers,
        )
        self.assertEqual(daily_resolved.status_code, 200, daily_resolved.text)
        resolved_row = next(r for r in daily_resolved.json() if r["employee_code"] == "EM013")
        self.assertEqual(resolved_row["exception_status"], "APPROVED")
        self.assertEqual(resolved_row["attendance_state"], "COMPLETE")
        self.assertEqual(resolved_row["checkout_status"], "ON_TIME")
        self.assertEqual(resolved_row["regular_minutes"], 540)
        self.assertEqual(resolved_row["overtime_minutes"], 0)
        self.assertEqual(resolved_row["payable_overtime_minutes"], 0)

        reopen_res = self.client.patch(
            f"/reports/attendance-exceptions/{exception_id}/reopen",
            headers=admin_headers,
        )
        self.assertEqual(reopen_res.status_code, 409, reopen_res.text)

        _FixedDateTime.fixed_now = None

    def test_exception_workflow_reject_requires_note_and_writes_audit_trail(self) -> None:
        admin = self._create_user(email="admin_exception_flow@example.com", password="admin123", role="ADMIN")
        user = self._create_user(email="exception_flow_user@example.com", password="user123", role="USER")
        employee = self._create_employee(code="EM099", full_name="Exception Flow User", user_id=user.id)
        self._create_rule(latitude=10.7769, longitude=106.7009, radius_m=300)

        token = self._login("exception_flow_user@example.com", "user123")
        headers = {"Authorization": f"Bearer {token}"}

        _FixedDateTime.fixed_now = datetime(2026, 3, 10, 1, 0, tzinfo=timezone.utc)
        with patch("app.api.attendance.datetime", _FixedDateTime):
            in_res = self.client.post(
                "/attendance/checkin",
                headers=headers,
                json={"lat": 10.7769, "lng": 106.7009},
            )
        self.assertEqual(in_res.status_code, 200, in_res.text)

        with SessionLocal() as db:
            source_checkin = (
                db.query(AttendanceLog)
                .filter(AttendanceLog.employee_id == employee.id, AttendanceLog.type == "IN")
                .order_by(AttendanceLog.id.desc())
                .first()
            )
            self.assertIsNotNone(source_checkin)
            source_checkin_id = source_checkin.id

        create_res = self.client.post(
            "/reports/attendance-exceptions/system",
            headers=self._system_headers(),
            json={
                "employee_id": employee.id,
                "source_checkin_log_id": source_checkin_id,
                "exception_type": "SUSPECTED_LOCATION_SPOOF",
                "note": "GPS risk detected by system",
            },
        )
        self.assertEqual(create_res.status_code, 200, create_res.text)
        created = create_res.json()
        exception_id = created["id"]
        self.assertEqual(created["status"], "PENDING_EMPLOYEE")

        invalid_approve_res = self.client.post(
            f"/reports/attendance-exceptions/{exception_id}/approve",
            headers={"Authorization": f"Bearer {self._login('admin_exception_flow@example.com', 'admin123')}"},
            json={"admin_note": "Cannot bypass employee"},
        )
        self.assertEqual(invalid_approve_res.status_code, 409, invalid_approve_res.text)

        submit_res = self.client.post(
            f"/reports/attendance-exceptions/{exception_id}/submit-explanation",
            headers=headers,
            json={"explanation": "Toi khong gia mao GPS, vui long kiem tra lai"},
        )
        self.assertEqual(submit_res.status_code, 200, submit_res.text)
        submitted = submit_res.json()
        self.assertEqual(submitted["status"], "PENDING_ADMIN")
        self.assertEqual(submitted["employee_explanation"], "Toi khong gia mao GPS, vui long kiem tra lai")

        admin_headers = {"Authorization": f"Bearer {self._login('admin_exception_flow@example.com', 'admin123')}"}
        reject_missing_note_res = self.client.post(
            f"/reports/attendance-exceptions/{exception_id}/reject",
            headers=admin_headers,
            json={"admin_note": ""},
        )
        self.assertEqual(reject_missing_note_res.status_code, 422, reject_missing_note_res.text)

        reject_res = self.client.post(
            f"/reports/attendance-exceptions/{exception_id}/reject",
            headers=admin_headers,
            json={"admin_note": "Khong du bang chung de chap nhan giai trinh"},
        )
        self.assertEqual(reject_res.status_code, 200, reject_res.text)
        rejected = reject_res.json()
        self.assertEqual(rejected["status"], "REJECTED")
        self.assertEqual(rejected["admin_note"], "Khong du bang chung de chap nhan giai trinh")
        self.assertEqual(rejected["decided_by"], admin.id)
        self.assertTrue(rejected["timeline"])

        with SessionLocal() as db:
            audits = (
                db.query(AttendanceExceptionAudit)
                .filter(AttendanceExceptionAudit.exception_id == exception_id)
                .order_by(AttendanceExceptionAudit.id.asc())
                .all()
            )

        self.assertEqual(
            [audit.event_type for audit in audits],
            [
                "exception_detected",
                "employee_explanation_submitted",
                "admin_rejected",
            ],
        )
        self.assertEqual(audits[0].next_status, "PENDING_EMPLOYEE")
        self.assertEqual(audits[1].previous_status, "PENDING_EMPLOYEE")
        self.assertEqual(audits[1].next_status, "PENDING_ADMIN")
        self.assertEqual(audits[2].previous_status, "PENDING_ADMIN")
        self.assertEqual(audits[2].next_status, "REJECTED")

        _FixedDateTime.fixed_now = None

    def test_exception_transition_blocks_invalid_pending_and_terminal_paths(self) -> None:
        admin = self._create_user(email="admin_transition_guard@example.com", password="admin123", role="ADMIN")
        user = self._create_user(email="transition_guard_user@example.com", password="user123", role="USER")
        employee = self._create_employee(code="EMTG1", full_name="Transition Guard User", user_id=user.id)
        admin_headers = {"Authorization": f"Bearer {create_access_token({'sub': str(admin.id), 'role': admin.role})}"}
        user_headers = {"Authorization": f"Bearer {create_access_token({'sub': str(user.id), 'role': user.role})}"}

        def create_checkin_log(work_date_value: date, happened_at: datetime) -> int:
            with SessionLocal() as db:
                log = AttendanceLog(
                    employee_id=employee.id,
                    type="IN",
                    time=happened_at,
                    work_date=work_date_value,
                    lat=10.7769,
                    lng=106.7009,
                    is_out_of_range=False,
                    punctuality_status="ON_TIME",
                )
                db.add(log)
                db.commit()
                db.refresh(log)
                return log.id

        def create_system_exception(source_log_id: int, exception_type: str, note: str) -> dict:
            res = self.client.post(
                "/reports/attendance-exceptions/system",
                headers=self._system_headers(),
                json={
                    "employee_id": employee.id,
                    "source_checkin_log_id": source_log_id,
                    "exception_type": exception_type,
                    "note": note,
                    "detected_at": "2026-03-20T00:00:00+00:00",
                    "expires_at": "2026-03-23T00:00:00+00:00",
                },
            )
            self.assertEqual(res.status_code, 200, res.text)
            return res.json()

        pending_employee = create_system_exception(
            create_checkin_log(date(2026, 3, 20), datetime(2026, 3, 20, 1, 0, tzinfo=timezone.utc)),
            "SUSPECTED_LOCATION_SPOOF",
            "GPS risk requires employee explanation",
        )
        self.assertEqual(pending_employee["status"], "PENDING_EMPLOYEE")
        pending_employee_id = pending_employee["id"]

        invalid_reject_res = self.client.post(
            f"/reports/attendance-exceptions/{pending_employee_id}/reject",
            headers=admin_headers,
            json={"admin_note": "Cannot reject before employee explanation"},
        )
        self.assertEqual(invalid_reject_res.status_code, 409, invalid_reject_res.text)

        submit_res = self.client.post(
            f"/reports/attendance-exceptions/{pending_employee_id}/submit-explanation",
            headers=user_headers,
            json={"explanation": "Employee explanation for GPS risk"},
        )
        self.assertEqual(submit_res.status_code, 200, submit_res.text)
        self.assertEqual(submit_res.json()["status"], "PENDING_ADMIN")

        approve_res = self.client.post(
            f"/reports/attendance-exceptions/{pending_employee_id}/approve",
            headers=admin_headers,
            json={"admin_note": "Accepted explanation"},
        )
        self.assertEqual(approve_res.status_code, 200, approve_res.text)
        self.assertEqual(approve_res.json()["status"], "APPROVED")

        submit_after_approved_res = self.client.post(
            f"/reports/attendance-exceptions/{pending_employee_id}/submit-explanation",
            headers=user_headers,
            json={"explanation": "Late edit after approval"},
        )
        self.assertEqual(submit_after_approved_res.status_code, 409, submit_after_approved_res.text)

        pending_admin = create_system_exception(
            create_checkin_log(date(2026, 3, 21), datetime(2026, 3, 21, 1, 0, tzinfo=timezone.utc)),
            "LARGE_TIME_DEVIATION",
            "Large time deviation can go directly to admin",
        )
        self.assertEqual(pending_admin["status"], "PENDING_ADMIN")
        pending_admin_id = pending_admin["id"]

        reject_res = self.client.post(
            f"/reports/attendance-exceptions/{pending_admin_id}/reject",
            headers=admin_headers,
            json={"admin_note": "Rejected direct-admin exception"},
        )
        self.assertEqual(reject_res.status_code, 200, reject_res.text)
        self.assertEqual(reject_res.json()["status"], "REJECTED")

        submit_after_rejected_res = self.client.post(
            f"/reports/attendance-exceptions/{pending_admin_id}/submit-explanation",
            headers=user_headers,
            json={"explanation": "Late edit after rejection"},
        )
        self.assertEqual(submit_after_rejected_res.status_code, 409, submit_after_rejected_res.text)

        expiring = create_system_exception(
            create_checkin_log(date(2026, 3, 22), datetime(2026, 3, 22, 1, 0, tzinfo=timezone.utc)),
            "SUSPECTED_LOCATION_SPOOF",
            "GPS risk will expire",
        )
        expire_res = self.client.post(
            f"/reports/attendance-exceptions/{expiring['id']}/expire",
            headers=self._system_headers(),
        )
        self.assertEqual(expire_res.status_code, 200, expire_res.text)
        self.assertEqual(expire_res.json()["status"], "EXPIRED")

        submit_after_expired_res = self.client.post(
            f"/reports/attendance-exceptions/{expiring['id']}/submit-explanation",
            headers=user_headers,
            json={"explanation": "Late edit after expiration"},
        )
        self.assertEqual(submit_after_expired_res.status_code, 409, submit_after_expired_res.text)

    def test_exception_workflow_notifications_for_action_endpoints(self) -> None:
        admin = self._create_user(email="admin_notify@example.com", password="admin123", role="ADMIN")
        user = self._create_user(email="notify_user@example.com", password="user123", role="USER")
        employee = self._create_employee(code="EMN01", full_name="Notify User", user_id=user.id)
        admin_headers = {"Authorization": f"Bearer {create_access_token({'sub': str(admin.id), 'role': admin.role})}"}
        user_headers = {"Authorization": f"Bearer {create_access_token({'sub': str(user.id), 'role': user.role})}"}

        first_log = self._create_attendance_log(
            employee_id=employee.id,
            work_date_value=date(2026, 3, 12),
            happened_at=datetime(2026, 3, 12, 1, 0, tzinfo=timezone.utc),
        )
        second_log = self._create_attendance_log(
            employee_id=employee.id,
            work_date_value=date(2026, 3, 13),
            happened_at=datetime(2026, 3, 13, 1, 0, tzinfo=timezone.utc),
        )
        third_log = self._create_attendance_log(
            employee_id=employee.id,
            work_date_value=date(2026, 3, 14),
            happened_at=datetime(2026, 3, 14, 1, 0, tzinfo=timezone.utc),
        )

        spy_mail = _SpyMailSender()
        with patch("app.services.attendance_exception_notifications.get_mail_sender", return_value=spy_mail):
            create_res = self.client.post(
                "/reports/attendance-exceptions/system",
                headers=self._system_headers(),
                json={
                    "employee_id": employee.id,
                    "source_checkin_log_id": first_log.id,
                    "exception_type": "SUSPECTED_LOCATION_SPOOF",
                    "note": "GPS risk detected",
                },
            )
            self.assertEqual(create_res.status_code, 200, create_res.text)
            exception_id = create_res.json()["id"]
            self.assertEqual(spy_mail.sent[-1].event_type, "exception_detected_employee")
            self.assertEqual(spy_mail.sent[-1].to_email, user.email)

            sent_count_before_invalid = len(spy_mail.sent)
            invalid_approve_res = self.client.post(
                f"/reports/attendance-exceptions/{exception_id}/approve",
                headers=admin_headers,
                json={"admin_note": "Cannot bypass employee"},
            )
            self.assertEqual(invalid_approve_res.status_code, 409, invalid_approve_res.text)
            self.assertEqual(len(spy_mail.sent), sent_count_before_invalid)

            submit_res = self.client.post(
                f"/reports/attendance-exceptions/{exception_id}/submit-explanation",
                headers=user_headers,
                json={"explanation": "Can admin kiem tra lai"},
            )
            self.assertEqual(submit_res.status_code, 200, submit_res.text)
            self.assertEqual(spy_mail.sent[-1].event_type, "exception_submitted_admin")
            self.assertEqual(spy_mail.sent[-1].to_email, admin.email)

            approve_res = self.client.post(
                f"/reports/attendance-exceptions/{exception_id}/approve",
                headers=admin_headers,
                json={"admin_note": "Chap nhan giai trinh"},
            )
            self.assertEqual(approve_res.status_code, 200, approve_res.text)
            self.assertEqual(spy_mail.sent[-1].event_type, "exception_approved_employee")
            self.assertEqual(spy_mail.sent[-1].to_email, user.email)

            direct_admin_res = self.client.post(
                "/reports/attendance-exceptions/system",
                headers=self._system_headers(),
                json={
                    "employee_id": employee.id,
                    "source_checkin_log_id": second_log.id,
                    "exception_type": "LARGE_TIME_DEVIATION",
                    "note": "Large time deviation",
                },
            )
            self.assertEqual(direct_admin_res.status_code, 200, direct_admin_res.text)
            self.assertEqual(direct_admin_res.json()["status"], "PENDING_ADMIN")
            self.assertEqual(spy_mail.sent[-1].event_type, "exception_detected_admin")
            self.assertEqual(spy_mail.sent[-1].to_email, admin.email)

            reject_create_res = self.client.post(
                "/reports/attendance-exceptions/system",
                headers=self._system_headers(),
                json={
                    "employee_id": employee.id,
                    "source_checkin_log_id": third_log.id,
                    "exception_type": "SUSPECTED_LOCATION_SPOOF",
                    "note": "GPS risk detected again",
                },
            )
            self.assertEqual(reject_create_res.status_code, 200, reject_create_res.text)
            reject_exception_id = reject_create_res.json()["id"]
            reject_submit_res = self.client.post(
                f"/reports/attendance-exceptions/{reject_exception_id}/submit-explanation",
                headers=user_headers,
                json={"explanation": "Giai trinh khong du bang chung"},
            )
            self.assertEqual(reject_submit_res.status_code, 200, reject_submit_res.text)
            reject_res = self.client.post(
                f"/reports/attendance-exceptions/{reject_exception_id}/reject",
                headers=admin_headers,
                json={"admin_note": "Khong du bang chung"},
            )
            self.assertEqual(reject_res.status_code, 200, reject_res.text)
            self.assertEqual(spy_mail.sent[-1].event_type, "exception_rejected_employee")
            self.assertEqual(spy_mail.sent[-1].to_email, user.email)

        with SessionLocal() as db:
            notifications = (
                db.query(AttendanceExceptionNotification)
                .order_by(AttendanceExceptionNotification.id.asc())
                .all()
            )
        self.assertIn("exception_detected_employee", [item.event_type for item in notifications])
        self.assertIn("exception_submitted_admin", [item.event_type for item in notifications])
        self.assertIn("exception_approved_employee", [item.event_type for item in notifications])
        self.assertIn("exception_detected_admin", [item.event_type for item in notifications])
        self.assertIn("exception_rejected_employee", [item.event_type for item in notifications])
        self.assertTrue(all(item.status == "SENT" for item in notifications))

    def test_exception_jobs_expire_and_reminder_are_idempotent(self) -> None:
        user = self._create_user(email="job_notify_user@example.com", password="user123", role="USER")
        employee = self._create_employee(code="EMJ01", full_name="Job Notify User", user_id=user.id)
        now = datetime(2026, 3, 15, 1, 0, tzinfo=timezone.utc)
        reminder_log = self._create_attendance_log(
            employee_id=employee.id,
            work_date_value=date(2026, 3, 15),
            happened_at=now,
        )
        expired_log = self._create_attendance_log(
            employee_id=employee.id,
            work_date_value=date(2026, 3, 16),
            happened_at=now + timedelta(days=1),
        )
        approved_log = self._create_attendance_log(
            employee_id=employee.id,
            work_date_value=date(2026, 3, 17),
            happened_at=now + timedelta(days=2),
        )
        rejected_log = self._create_attendance_log(
            employee_id=employee.id,
            work_date_value=date(2026, 3, 18),
            happened_at=now + timedelta(days=3),
        )

        with SessionLocal() as db:
            reminder_exception = AttendanceException(
                employee_id=employee.id,
                source_checkin_log_id=reminder_log.id,
                exception_type="SUSPECTED_LOCATION_SPOOF",
                work_date=date(2026, 3, 15),
                status="PENDING_EMPLOYEE",
                detected_at=now - timedelta(days=2),
                expires_at=now + timedelta(hours=12),
            )
            expired_exception = AttendanceException(
                employee_id=employee.id,
                source_checkin_log_id=expired_log.id,
                exception_type="SUSPECTED_LOCATION_SPOOF",
                work_date=date(2026, 3, 16),
                status="PENDING_EMPLOYEE",
                detected_at=now - timedelta(days=4),
                expires_at=now - timedelta(minutes=1),
            )
            approved_exception = AttendanceException(
                employee_id=employee.id,
                source_checkin_log_id=approved_log.id,
                exception_type="SUSPECTED_LOCATION_SPOOF",
                work_date=date(2026, 3, 17),
                status="APPROVED",
                detected_at=now - timedelta(days=5),
                expires_at=now - timedelta(days=1),
            )
            rejected_exception = AttendanceException(
                employee_id=employee.id,
                source_checkin_log_id=rejected_log.id,
                exception_type="SUSPECTED_LOCATION_SPOOF",
                work_date=date(2026, 3, 18),
                status="REJECTED",
                detected_at=now - timedelta(days=5),
                expires_at=now - timedelta(days=1),
            )
            db.add_all([
                reminder_exception,
                expired_exception,
                approved_exception,
                rejected_exception,
            ])
            db.commit()
            db.refresh(reminder_exception)
            db.refresh(expired_exception)
            db.refresh(approved_exception)
            db.refresh(rejected_exception)
            reminder_exception_id = reminder_exception.id
            expired_exception_id = expired_exception.id
            approved_exception_id = approved_exception.id
            rejected_exception_id = rejected_exception.id

        spy_mail = _SpyMailSender()
        with patch("app.services.attendance_exception_notifications.get_mail_sender", return_value=spy_mail):
            with SessionLocal() as db:
                self.assertEqual(send_expire_reminders(db, now=now), 1)
            with SessionLocal() as db:
                self.assertEqual(send_expire_reminders(db, now=now), 0)
            with SessionLocal() as db:
                self.assertEqual(expire_overdue_exceptions(db, now=now), 1)
            with SessionLocal() as db:
                self.assertEqual(expire_overdue_exceptions(db, now=now), 0)

        self.assertEqual(
            [payload.event_type for payload in spy_mail.sent],
            [
                "exception_expire_reminder_employee",
                "exception_expired_employee",
            ],
        )

        with SessionLocal() as db:
            refreshed_reminder = db.query(AttendanceException).filter(AttendanceException.id == reminder_exception_id).first()
            refreshed_expired = db.query(AttendanceException).filter(AttendanceException.id == expired_exception_id).first()
            refreshed_approved = db.query(AttendanceException).filter(AttendanceException.id == approved_exception_id).first()
            refreshed_rejected = db.query(AttendanceException).filter(AttendanceException.id == rejected_exception_id).first()
            notifications = (
                db.query(AttendanceExceptionNotification)
                .order_by(AttendanceExceptionNotification.event_type.asc())
                .all()
            )
            expire_audit = (
                db.query(AttendanceExceptionAudit)
                .filter(
                    AttendanceExceptionAudit.exception_id == expired_exception_id,
                    AttendanceExceptionAudit.event_type == "system_expired",
                )
                .first()
            )

        self.assertEqual(refreshed_reminder.status, "PENDING_EMPLOYEE")
        self.assertEqual(refreshed_expired.status, "EXPIRED")
        self.assertEqual(refreshed_approved.status, "APPROVED")
        self.assertEqual(refreshed_rejected.status, "REJECTED")
        self.assertIsNotNone(expire_audit)
        self.assertEqual(
            sorted(item.event_type for item in notifications),
            [
                "exception_expire_reminder_employee",
                "exception_expired_employee",
            ],
        )
        self.assertTrue(all(item.status == "SENT" for item in notifications))

    def test_same_day_open_in_still_blocks_second_checkin(self) -> None:
        user = self._create_user(email="same_day_user@example.com", password="user123", role="USER")
        self._create_employee(code="EM012", full_name="Same Day User", user_id=user.id)
        self._create_rule(latitude=10.7769, longitude=106.7009, radius_m=300)

        token = self._login("same_day_user@example.com", "user123")
        headers = {"Authorization": f"Bearer {token}"}

        _FixedDateTime.fixed_now = datetime(2026, 3, 11, 2, 0, tzinfo=timezone.utc)
        with patch("app.api.attendance.datetime", _FixedDateTime):
            first_in = self.client.post(
                "/attendance/checkin",
                headers=headers,
                json={"lat": 10.7769, "lng": 106.7009},
            )
            second_in = self.client.post(
                "/attendance/checkin",
                headers=headers,
                json={"lat": 10.7769, "lng": 106.7009},
            )

        self.assertEqual(first_in.status_code, 200, first_in.text)
        self.assertEqual(second_in.status_code, 400, second_in.text)

        with SessionLocal() as db:
            emp = db.query(Employee).filter(Employee.user_id == user.id).first()
            self.assertIsNotNone(emp)
            count_exceptions = db.query(AttendanceException).filter(AttendanceException.employee_id == emp.id).count()
            self.assertEqual(count_exceptions, 0)

        _FixedDateTime.fixed_now = None
    def test_distance_consistency_warning_is_synced_between_daily_and_excel(self) -> None:
        admin = self._create_user(email="admin_distance_warn@example.com", password="admin123", role="ADMIN")
        user = self._create_user(email="distance_warn_user@example.com", password="user123", role="USER")

        self._create_employee(code="EM014", full_name="Distance Warning User", user_id=user.id)
        self._create_rule(latitude=10.7769, longitude=106.7009, radius_m=200)

        user_token = self._login("distance_warn_user@example.com", "user123")
        user_headers = {"Authorization": f"Bearer {user_token}"}

        in_res = self.client.post(
            "/attendance/checkin",
            headers=user_headers,
            json={"lat": 10.7769, "lng": 106.7009},
        )
        self.assertEqual(in_res.status_code, 200, in_res.text)

        out_res = self.client.post(
            "/attendance/checkout",
            headers=user_headers,
            json={"lat": 10.7769, "lng": 106.7009},
        )
        self.assertEqual(out_res.status_code, 200, out_res.text)

        with SessionLocal() as db:
            emp = db.query(Employee).filter(Employee.user_id == user.id).first()
            self.assertIsNotNone(emp)
            logs = db.query(AttendanceLog).filter(AttendanceLog.employee_id == emp.id).all()
            self.assertTrue(logs)
            for log in logs:
                log.distance_m = 350.0
                log.is_out_of_range = False
            db.commit()

        admin_token = self._login("admin_distance_warn@example.com", "admin123")
        admin_headers = {"Authorization": f"Bearer {admin_token}"}

        today = date.today().isoformat()
        daily_res = self.client.get(
            f"/attendance/report/daily?from_date={today}&to_date={today}",
            headers=admin_headers,
        )
        self.assertEqual(daily_res.status_code, 200, daily_res.text)

        daily_rows = [r for r in daily_res.json() if r["employee_code"] == "EM014"]
        self.assertTrue(daily_rows)
        daily_row = daily_rows[0]
        self.assertEqual(daily_row["out_of_range"], False)
        self.assertEqual(daily_row["radius_m"], 200)
        self.assertEqual(
            daily_row["distance_consistency_warning"],
            "IN_RANGE_DISTANCE_EXCEEDS_RADIUS",
        )

        excel_res = self.client.get(
            f"/reports/attendance.xlsx?from={today}&to={today}",
            headers=admin_headers,
        )
        self.assertEqual(excel_res.status_code, 200, excel_res.text)

        workbook = load_workbook(BytesIO(excel_res.content))
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[1]]
        header_index = {name: idx + 1 for idx, name in enumerate(headers)}

        found = False
        for row_idx in range(2, worksheet.max_row + 1):
            code = worksheet.cell(row=row_idx, column=header_index["employee_code"]).value
            if code != "EM014":
                continue
            found = True
            radius_value = worksheet.cell(row=row_idx, column=header_index["radius_m"]).value
            warning_value = worksheet.cell(row=row_idx, column=header_index["distance_consistency_warning"]).value
            self.assertEqual(radius_value, 200)
            self.assertEqual(warning_value, "IN_RANGE_DISTANCE_EXCEEDS_RADIUS")
            break

        self.assertTrue(found)
    def test_export_report_flow(self) -> None:
        admin = self._create_user(email="admin_report@example.com", password="admin123", role="ADMIN")
        user = self._create_user(email="user_report@example.com", password="user123", role="USER")

        report_group = self._create_group("REP", "Report Group")
        self._create_geofence(report_group.id, "Report Gate", 10.7769, 106.7009, 300)

        self._create_employee(code="AD001", full_name="Admin", user_id=admin.id)
        self._create_employee(code="EM002", full_name="User Report", user_id=user.id, group_id=report_group.id)
        self._create_rule()

        user_token = self._login("user_report@example.com", "user123")
        user_headers = {"Authorization": f"Bearer {user_token}"}

        in_res = self.client.post(
            "/attendance/checkin",
            headers=user_headers,
            json={"lat": 10.7769, "lng": 106.7009},
        )
        self.assertEqual(in_res.status_code, 200, in_res.text)

        out_res = self.client.post(
            "/attendance/checkout",
            headers=user_headers,
            json={"lat": 10.7769, "lng": 106.7009},
        )
        self.assertEqual(out_res.status_code, 200, out_res.text)

        admin_token = self._login("admin_report@example.com", "admin123")
        today = date.today().isoformat()

        report_res = self.client.get(
            f"/reports/attendance.xlsx?from={today}&to={today}&group_id={report_group.id}",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        self.assertEqual(report_res.status_code, 200, report_res.text)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            report_res.headers.get("content-type", ""),
        )
        self.assertTrue(report_res.content.startswith(b"PK"))

        workbook = load_workbook(BytesIO(report_res.content))
        worksheet = workbook.active

        headers = [cell.value for cell in worksheet[1]]
        for required_header in ("group_code", "group_name", "matched_geofence", "geofence_source", "radius_m", "distance_consistency_warning", "payable_overtime_minutes"):
            self.assertIn(required_header, headers)

        header_index = {name: idx + 1 for idx, name in enumerate(headers)}
        self.assertGreaterEqual(worksheet.max_row, 2)

        self.assertEqual(worksheet.cell(row=2, column=header_index["group_code"]).value, "REP")
        self.assertEqual(worksheet.cell(row=2, column=header_index["group_name"]).value, "Report Group")
        self.assertEqual(worksheet.cell(row=2, column=header_index["matched_geofence"]).value, "Report Gate")
        self.assertEqual(worksheet.cell(row=2, column=header_index["geofence_source"]).value, "GROUP")
        self.assertEqual(worksheet.cell(row=2, column=header_index["out_of_range"]).value, "IN_RANGE")



    def test_geofence_radius_small_vs_large_affects_out_of_range(self) -> None:
        user = self._create_user(email="radius_compare_user@example.com", password="user123", role="USER")
        self._create_employee(code="EM015", full_name="Radius Compare", user_id=user.id)

        self._create_rule(latitude=10.7769, longitude=106.7009, radius_m=20)

        token = self._login("radius_compare_user@example.com", "user123")
        headers = {"Authorization": f"Bearer {token}"}

        _FixedDateTime.fixed_now = datetime(2026, 3, 10, 2, 0, tzinfo=timezone.utc)
        with patch("app.api.attendance.datetime", _FixedDateTime):
            day1_in = self.client.post(
                "/attendance/checkin",
                headers=headers,
                json={"lat": 10.7774, "lng": 106.7009},
            )
        self.assertEqual(day1_in.status_code, 200, day1_in.text)
        self.assertTrue(day1_in.json()["log"]["is_out_of_range"])

        _FixedDateTime.fixed_now = datetime(2026, 3, 10, 3, 0, tzinfo=timezone.utc)
        with patch("app.api.attendance.datetime", _FixedDateTime):
            day1_out = self.client.post(
                "/attendance/checkout",
                headers=headers,
                json={"lat": 10.7774, "lng": 106.7009},
            )
        self.assertEqual(day1_out.status_code, 200, day1_out.text)
        self.assertTrue(day1_out.json()["log"]["is_out_of_range"])

        with SessionLocal() as db:
            active_rule = db.query(CheckinRule).filter(CheckinRule.active.is_(True)).first()
            self.assertIsNotNone(active_rule)
            assert active_rule is not None
            active_rule.radius_m = 500
            db.commit()

        _FixedDateTime.fixed_now = datetime(2026, 3, 11, 2, 0, tzinfo=timezone.utc)
        with patch("app.api.attendance.datetime", _FixedDateTime):
            day2_in = self.client.post(
                "/attendance/checkin",
                headers=headers,
                json={"lat": 10.7774, "lng": 106.7009},
            )
        self.assertEqual(day2_in.status_code, 200, day2_in.text)
        self.assertFalse(day2_in.json()["log"]["is_out_of_range"])

        _FixedDateTime.fixed_now = None

    def test_cross_day_cutoff_daily_and_excel_are_consistent(self) -> None:
        admin = self._create_user(email="admin_cross_day_consistency@example.com", password="admin123", role="ADMIN")
        user = self._create_user(email="cross_day_consistency_user@example.com", password="user123", role="USER")

        self._create_employee(code="EM016", full_name="Cross Day Consistency", user_id=user.id)
        self._create_rule(
            latitude=10.7769,
            longitude=106.7009,
            radius_m=300,
            start_time=time(8, 0),
            grace_minutes=30,
            end_time=time(17, 0),
            checkout_grace_minutes=0,
            cross_day_cutoff_minutes=240,
        )

        user_token = self._login("cross_day_consistency_user@example.com", "user123")
        user_headers = {"Authorization": f"Bearer {user_token}"}

        _FixedDateTime.fixed_now = datetime(2026, 3, 10, 1, 0, tzinfo=timezone.utc)
        with patch("app.api.attendance.datetime", _FixedDateTime):
            in_res = self.client.post(
                "/attendance/checkin",
                headers=user_headers,
                json={"lat": 10.7769, "lng": 106.7009},
            )
        self.assertEqual(in_res.status_code, 200, in_res.text)

        _FixedDateTime.fixed_now = datetime(2026, 3, 10, 19, 30, tzinfo=timezone.utc)
        with patch("app.api.attendance.datetime", _FixedDateTime):
            out_res = self.client.post(
                "/attendance/checkout",
                headers=user_headers,
                json={"lat": 10.7769, "lng": 106.7009},
            )
        self.assertEqual(out_res.status_code, 200, out_res.text)

        admin_token = self._login("admin_cross_day_consistency@example.com", "admin123")
        admin_headers = {"Authorization": f"Bearer {admin_token}"}

        daily_res = self.client.get(
            "/attendance/report/daily?from_date=2026-03-10&to_date=2026-03-10",
            headers=admin_headers,
        )
        self.assertEqual(daily_res.status_code, 200, daily_res.text)
        daily_row = next(r for r in daily_res.json() if r["employee_code"] == "EM016")

        excel_res = self.client.get(
            "/reports/attendance.xlsx?from=2026-03-10&to=2026-03-10",
            headers=admin_headers,
        )
        self.assertEqual(excel_res.status_code, 200, excel_res.text)

        workbook = load_workbook(BytesIO(excel_res.content))
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[1]]
        header_index = {name: idx + 1 for idx, name in enumerate(headers)}

        excel_row_idx = None
        for row_idx in range(2, worksheet.max_row + 1):
            if worksheet.cell(row=row_idx, column=header_index["employee_code"]).value == "EM016":
                excel_row_idx = row_idx
                break

        self.assertIsNotNone(excel_row_idx)
        assert excel_row_idx is not None

        self.assertEqual(worksheet.cell(row=excel_row_idx, column=header_index["regular_minutes"]).value, daily_row["regular_minutes"])
        self.assertEqual(worksheet.cell(row=excel_row_idx, column=header_index["overtime_minutes"]).value, daily_row["overtime_minutes"])
        self.assertEqual(worksheet.cell(row=excel_row_idx, column=header_index["payable_overtime_minutes"]).value, daily_row["payable_overtime_minutes"])
        self.assertEqual(
            worksheet.cell(row=excel_row_idx, column=header_index["overtime_cross_day"]).value,
            "YES" if daily_row["overtime_cross_day"] else "NO",
        )

        _FixedDateTime.fixed_now = None

    def test_race_condition_checkout_concurrent_only_one_out_log(self) -> None:
        user = self._create_user(email="race_checkout_user@example.com", password="user123", role="USER")
        self._create_employee(code="EM017", full_name="Race Checkout", user_id=user.id)
        self._create_rule(latitude=10.7769, longitude=106.7009, radius_m=300)

        token = self._login("race_checkout_user@example.com", "user123")
        headers = {"Authorization": f"Bearer {token}"}

        in_res = self.client.post(
            "/attendance/checkin",
            headers=headers,
            json={"lat": 10.7769, "lng": 106.7009},
        )
        self.assertEqual(in_res.status_code, 200, in_res.text)

        barrier = threading.Barrier(2)
        results: list[tuple[str, int | str]] = []
        lock = threading.Lock()

        def _worker_checkout() -> None:
            with SessionLocal() as db:
                user_obj = db.query(User).filter(User.id == user.id).first()
                assert user_obj is not None
                payload = LocationRequest(lat=10.7769, lng=106.7009)
                barrier.wait()
                try:
                    attendance_api.checkout(payload, db=db, user=user_obj)
                    item: tuple[str, int | str] = ("ok", 200)
                except HTTPException as exc:
                    item = ("http", exc.status_code)
                except Exception as exc:  # pragma: no cover - should not happen
                    item = ("error", str(exc))
            with lock:
                results.append(item)

        t1 = threading.Thread(target=_worker_checkout)
        t2 = threading.Thread(target=_worker_checkout)
        t1.start()
        t2.start()
        t1.join()
        t2.join()

        self.assertEqual(len(results), 2)

        ok_count = sum(1 for kind, _ in results if kind == "ok")
        http_400_count = sum(1 for kind, code in results if kind == "http" and code == 400)
        errors = [value for kind, value in results if kind == "error"]

        self.assertEqual(errors, [])
        self.assertEqual(ok_count, 1)
        self.assertEqual(http_400_count, 1)

        with SessionLocal() as db:
            emp = db.query(Employee).filter(Employee.user_id == user.id).first()
            self.assertIsNotNone(emp)
            out_logs = (
                db.query(AttendanceLog)
                .filter(
                    AttendanceLog.employee_id == emp.id,
                    AttendanceLog.type == "OUT",
                )
                .all()
            )
            self.assertEqual(len(out_logs), 1)
if __name__ == "__main__":
    unittest.main()




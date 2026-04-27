from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO
import json

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, case, func, or_
from sqlalchemy.orm import Session, aliased

from app.core.db import get_db
from app.core.deps import get_current_user, require_admin, require_exception_workflow_system
from app.models import AttendanceException, AttendanceExceptionAudit, AttendanceExceptionNotification, AttendanceLog, CheckinRule, Employee, ExceptionPolicy, Group, GroupGeofence, LeaveRequest, PublicHoliday, User
from app.schemas.attendance import (
    AttendanceExceptionApproveRequest,
    AttendanceExceptionAuditResponse,
    AttendanceExceptionCreateRequest,
    AttendanceExceptionDetailResponse,
    AttendanceExceptionExtendDeadlineRequest,
    AttendanceExceptionRejectRequest,
    AttendanceExceptionReportResponse,
    AttendanceExceptionResolveRequest,
    AttendanceExceptionSubmitExplanationRequest,
)
from app.services.attendance_exception_workflow import (
    APPROVED,
    EXPIRED,
    PENDING_ADMIN,
    PENDING_EMPLOYEE,
    REJECTED,
    auto_expire_overdue,
    default_exception_status_for_type,
    build_exception_status_filter_values,
    ensure_allowed_exception_transition,
    get_deadline_hours,
    get_effective_deadline,
    is_pending_exception_status,
    is_pending_timesheet_exception,
    normalize_exception_status,
)
from app.services.attendance_exception_audit import record_attendance_exception_audit
from app.services.attendance_exception_notifications import (
    build_exception_notification_mail,
    create_exception_notification_record,
    send_exception_notification_background,
)
from app.services.attendance_time import (
    DEFAULT_CROSS_DAY_CUTOFF_MINUTES,
    classify_checkout_status,
    normalize_utc,
    split_regular_overtime_minutes,
    to_vn_time,
    work_date_cutoff_utc,
)
from app.services.report_consistency import (
    compute_distance_consistency_warning,
    load_group_geofence_radius_maps,
    resolve_reference_radius_m,
)

router = APIRouter(prefix="/reports", tags=["reports"])
VN_TZ = timezone(timedelta(hours=7))
UTC_TZ = timezone.utc


def _deserialize_risk_flags(raw: str | None) -> list[str]:
    if not raw:
        return []
    try:
        parsed = json.loads(raw)
    except (TypeError, ValueError, json.JSONDecodeError):
        return []
    if isinstance(parsed, list):
        return [str(x) for x in parsed if str(x).strip()]
    return []


def _deserialize_metadata_json(raw: str | None) -> dict[str, object]:
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
    except (TypeError, ValueError, json.JSONDecodeError):
        return {}
    return parsed if isinstance(parsed, dict) else {}


def _normalize_exception_type(value: str) -> str:
    if value == "GPS_RISK":
        return "SUSPECTED_LOCATION_SPOOF"
    return value


def _rank_to_punctuality(rank_value) -> str | None:
    if rank_value is None:
        return None
    mapping = {1: "EARLY", 2: "ON_TIME", 3: "LATE"}
    return mapping.get(int(rank_value))


def _rank_to_geofence_source(rank_value) -> str | None:
    if rank_value is None:
        return None
    mapping = {1: "GROUP", 2: "SYSTEM_FALLBACK"}
    return mapping.get(int(rank_value))


def _derive_daily_status(
    checkin_time: datetime | None,
    checkout_time: datetime | None,
    checkin_rank,
    checkout_rank,
    checkout_raw_status: str | None = None,
) -> tuple[str | None, str | None, str]:
    checkin_status = _rank_to_punctuality(checkin_rank)
    checkout_status = _rank_to_punctuality(checkout_rank)

    if checkout_raw_status in {"SYSTEM_AUTO", "MISSING_PUNCH"}:
        checkout_status = checkout_raw_status

    if checkin_time is None and checkout_time is None:
        return "NO_CHECKIN", "NO_CHECKOUT", "ABSENT"

    if checkin_time is None:
        return "NO_CHECKIN", checkout_status or "NO_CHECKOUT", "MISSING_CHECKIN_ANOMALY"

    if checkout_time is None:
        return checkin_status, "NO_CHECKOUT", "MISSED_CHECKOUT"

    return checkin_status, checkout_status, "COMPLETE"


def _work_date_expr(db: Session):
    dialect = db.bind.dialect.name if db.bind is not None else ""
    if dialect == "postgresql":
        legacy_expr = func.date(func.timezone("Asia/Ho_Chi_Minh", AttendanceLog.time))
    else:
        legacy_expr = func.date(AttendanceLog.time)
    return func.coalesce(AttendanceLog.work_date, legacy_expr)


def _fetch_daily_report_rows(
    db: Session,
    from_date: date | None,
    to_date: date | None,
    employee_id: int | None,
    group_id: int | None,
):
    work_date_expr = _work_date_expr(db)

    checkin_time_expr = func.min(case((AttendanceLog.type == "IN", AttendanceLog.time), else_=None)).label("checkin_time")
    checkout_time_expr = func.max(case((AttendanceLog.type == "OUT", AttendanceLog.time), else_=None)).label("checkout_time")

    punctuality_rank_expr = func.min(
        case(
            (
                AttendanceLog.type == "IN",
                case(
                    (AttendanceLog.punctuality_status == "EARLY", 1),
                    (AttendanceLog.punctuality_status == "ON_TIME", 2),
                    (AttendanceLog.punctuality_status == "LATE", 3),
                    else_=None,
                ),
            ),
            else_=None,
        )
    ).label("punctuality_rank")

    checkout_rank_expr = func.max(
        case(
            (
                AttendanceLog.type == "OUT",
                case(
                    (AttendanceLog.checkout_status == "EARLY", 1),
                    (AttendanceLog.checkout_status == "ON_TIME", 2),
                    (AttendanceLog.checkout_status == "LATE", 3),
                    else_=None,
                ),
            ),
            else_=None,
        )
    ).label("checkout_rank")

    checkout_raw_status_expr = func.max(
        case((AttendanceLog.type == "OUT", AttendanceLog.checkout_status), else_=None)
    ).label("checkout_raw_status")


    checkin_matched_geofence_expr = func.max(
        case((AttendanceLog.type == "IN", AttendanceLog.matched_geofence_name), else_=None)
    ).label("checkin_matched_geofence")
    checkout_matched_geofence_expr = func.max(
        case((AttendanceLog.type == "OUT", AttendanceLog.matched_geofence_name), else_=None)
    ).label("checkout_matched_geofence")

    geofence_source_rank_expr = func.min(
        case(
            (AttendanceLog.geofence_source == "GROUP", 1),
            (AttendanceLog.geofence_source == "SYSTEM_FALLBACK", 2),
            else_=None,
        )
    ).label("geofence_source_rank")
    out_of_range_expr = func.bool_or(AttendanceLog.is_out_of_range).label("out_of_range")
    avg_distance_expr = func.avg(AttendanceLog.distance_m).label("avg_distance_m")
    max_distance_expr = func.max(AttendanceLog.distance_m).label("max_distance_m")
    shift_start_expr = func.max(case((AttendanceLog.type == "IN", AttendanceLog.snapshot_start_time), else_=None)).label("shift_start")
    shift_end_expr = func.max(case((AttendanceLog.type == "IN", AttendanceLog.snapshot_end_time), else_=None)).label("shift_end")

    checkin_lat_expr = func.max(
        case((AttendanceLog.type == "IN", AttendanceLog.lat), else_=None)
    ).label("checkin_lat")
    checkin_lng_expr = func.max(
        case((AttendanceLog.type == "IN", AttendanceLog.lng), else_=None)
    ).label("checkin_lng")
    checkout_lat_expr = func.max(
        case((AttendanceLog.type == "OUT", AttendanceLog.lat), else_=None)
    ).label("checkout_lat")
    checkout_lng_expr = func.max(
        case((AttendanceLog.type == "OUT", AttendanceLog.lng), else_=None)
    ).label("checkout_lng")

    q = (
        db.query(
            work_date_expr.label("work_date"),
            Employee.id.label("employee_id"),
            Employee.code.label("employee_code"),
            Employee.full_name.label("full_name"),
            Group.id.label("group_id"),
            Group.code.label("group_code"),
            Group.name.label("group_name"),
            checkin_time_expr,
            checkout_time_expr,
            punctuality_rank_expr,
            checkout_rank_expr,
            checkout_raw_status_expr,
            checkin_matched_geofence_expr,
            checkout_matched_geofence_expr,
            geofence_source_rank_expr,
            out_of_range_expr,
            avg_distance_expr,
            max_distance_expr,
            shift_start_expr,
            shift_end_expr,
            checkin_lat_expr,
            checkin_lng_expr,
            checkout_lat_expr,
            checkout_lng_expr,
        )
        .join(Employee, Employee.id == AttendanceLog.employee_id)
        .outerjoin(Group, Group.id == Employee.group_id)
    )

    if employee_id:
        q = q.filter(AttendanceLog.employee_id == employee_id)
    if group_id:
        q = q.filter(Employee.group_id == group_id)
    if from_date:
        q = q.filter(work_date_expr >= from_date)
    if to_date:
        q = q.filter(work_date_expr <= to_date)

    return (
        q.group_by(work_date_expr, Employee.id, Employee.code, Employee.full_name, Group.id, Group.code, Group.name)
        .order_by(work_date_expr.asc(), Employee.code.asc())
        .all()
    )


def _build_exception_map(
    db: Session,
    from_date: date | None,
    to_date: date | None,
    employee_id: int | None,
    group_id: int | None,
) -> dict[tuple[int, date], tuple[str, str]]:
    q = db.query(
        AttendanceException.employee_id,
        AttendanceException.work_date,
        AttendanceException.status,
        AttendanceException.exception_type,
    ).join(Employee, Employee.id == AttendanceException.employee_id)

    if employee_id:
        q = q.filter(AttendanceException.employee_id == employee_id)
    if group_id:
        q = q.filter(Employee.group_id == group_id)
    if from_date:
        q = q.filter(AttendanceException.work_date >= from_date)
    if to_date:
        q = q.filter(AttendanceException.work_date <= to_date)

    status_map: dict[tuple[int, date], tuple[str, str]] = {}
    for row in q.all():
        key = (row.employee_id, row.work_date)
        current = status_map.get(key)
        normalized_status = normalize_exception_status(row.status)
        if current is None or not is_pending_exception_status(current[0]):
            status_map[key] = (normalized_status, row.exception_type)
    return status_map


def _apply_exception_to_attendance_state(
    attendance_state: str,
    exception_status: str | None,
    exception_type: str | None,
) -> str:
    if is_pending_timesheet_exception(exception_status, exception_type):
        return "PENDING_TIMESHEET"
    return attendance_state



def _compute_payable_overtime_minutes(
    overtime_minutes: int | None,
    exception_status: str | None,
    exception_type: str | None,
) -> int | None:
    if overtime_minutes is None:
        return None
    if is_pending_timesheet_exception(exception_status, exception_type):
        return 0
    return overtime_minutes


def _find_checkout_log_for_exception(db: Session, exception: AttendanceException, source_checkin: AttendanceLog) -> AttendanceLog | None:
    return (
        db.query(AttendanceLog)
        .filter(
            AttendanceLog.employee_id == exception.employee_id,
            AttendanceLog.work_date == exception.work_date,
            AttendanceLog.type == "OUT",
            AttendanceLog.time >= source_checkin.time,
        )
        .order_by(AttendanceLog.time.asc())
        .first()
    )


def _upsert_checkout_log_from_resolution(
    db: Session,
    exception: AttendanceException,
    source_checkin: AttendanceLog,
    actual_checkout_utc: datetime,
) -> AttendanceLog:
    checkout_log = _find_checkout_log_for_exception(db, exception, source_checkin)

    shift_end = source_checkin.snapshot_end_time or time(17, 0)
    checkout_grace = (
        source_checkin.snapshot_checkout_grace_minutes
        if source_checkin.snapshot_checkout_grace_minutes is not None
        else 0
    )
    checkout_status = classify_checkout_status(
        actual_checkout_utc,
        shift_end,
        checkout_grace,
        work_date=exception.work_date,
    )

    if checkout_log is None:
        checkout_log = AttendanceLog(
            employee_id=exception.employee_id,
            type="OUT",
            time=actual_checkout_utc,
            work_date=exception.work_date,
            lat=source_checkin.lat,
            lng=source_checkin.lng,
            distance_m=source_checkin.distance_m,
            is_out_of_range=source_checkin.is_out_of_range,
            checkout_status=checkout_status,
            matched_geofence_name=source_checkin.matched_geofence_name,
            geofence_source=source_checkin.geofence_source,
            fallback_reason=source_checkin.fallback_reason,
            snapshot_start_time=source_checkin.snapshot_start_time,
            snapshot_end_time=source_checkin.snapshot_end_time,
            snapshot_grace_minutes=source_checkin.snapshot_grace_minutes,
            snapshot_checkout_grace_minutes=source_checkin.snapshot_checkout_grace_minutes,
            snapshot_cutoff_minutes=source_checkin.snapshot_cutoff_minutes,
            time_rule_source=source_checkin.time_rule_source,
            time_rule_fallback_reason=source_checkin.time_rule_fallback_reason,
            address_text="RESOLVED_BY_ADMIN",
        )
        db.add(checkout_log)
    else:
        checkout_log.time = actual_checkout_utc
        checkout_log.checkout_status = checkout_status
        checkout_log.address_text = "RESOLVED_BY_ADMIN"

    return checkout_log


def _revert_checkout_log_for_reopen(
    db: Session,
    exception: AttendanceException,
    source_checkin: AttendanceLog,
) -> None:
    checkout_log = _find_checkout_log_for_exception(db, exception, source_checkin)
    if checkout_log is None:
        return

    if exception.exception_type == "AUTO_CLOSED":
        cutoff_minutes = source_checkin.snapshot_cutoff_minutes or DEFAULT_CROSS_DAY_CUTOFF_MINUTES
        checkout_log.time = work_date_cutoff_utc(exception.work_date, cutoff_minutes)
        checkout_log.checkout_status = "SYSTEM_AUTO"
        checkout_log.address_text = "AUTO_CLOSED_AT_CUTOFF"
        return

    if exception.exception_type == "MISSED_CHECKOUT" and checkout_log.address_text == "RESOLVED_BY_ADMIN":
        db.delete(checkout_log)


def _can_submit_explanation(status: str | None) -> bool:
    return normalize_exception_status(status) == PENDING_EMPLOYEE


def _can_admin_decide(status: str | None) -> bool:
    return normalize_exception_status(status) == PENDING_ADMIN


def _can_expire(status: str | None, expires_at: datetime | None) -> bool:
    normalized = normalize_exception_status(status)
    if normalized != PENDING_EMPLOYEE or expires_at is None:
        return False
    return normalize_utc(expires_at) <= datetime.now(timezone.utc)


def _build_exception_timeline(db: Session, exception_id: int) -> list[AttendanceExceptionAuditResponse]:
    rows = (
        db.query(AttendanceExceptionAudit)
        .filter(AttendanceExceptionAudit.exception_id == exception_id)
        .order_by(AttendanceExceptionAudit.created_at.asc(), AttendanceExceptionAudit.id.asc())
        .all()
    )
    return [
        AttendanceExceptionAuditResponse(
            id=row.id,
            event_type=row.event_type,
            previous_status=normalize_exception_status(row.previous_status),
            next_status=normalize_exception_status(row.next_status),
            actor_type=row.actor_type,
            actor_id=row.actor_id,
            actor_email=row.actor_email,
            metadata=_deserialize_metadata_json(row.metadata_json),
            created_at=row.created_at,
        )
        for row in rows
    ]


def _build_exception_response(db: Session, exception_id: int) -> AttendanceExceptionDetailResponse:
    resolver = aliased(User)
    decider = aliased(User)
    row = (
        db.query(
            AttendanceException.id.label("id"),
            AttendanceException.employee_id.label("employee_id"),
            Employee.code.label("employee_code"),
            Employee.full_name.label("full_name"),
            Group.id.label("group_id"),
            Group.code.label("group_code"),
            Group.name.label("group_name"),
            AttendanceException.work_date.label("work_date"),
            AttendanceException.exception_type.label("exception_type"),
            AttendanceException.status.label("status"),
            AttendanceException.note.label("note"),
            AttendanceException.resolved_note.label("resolved_note"),
            AttendanceException.source_checkin_log_id.label("source_checkin_log_id"),
            AttendanceLog.time.label("source_checkin_time"),
            AttendanceLog.risk_score.label("risk_score"),
            AttendanceLog.risk_level.label("risk_level"),
            AttendanceLog.risk_flags.label("risk_flags"),
            AttendanceLog.risk_policy_version.label("risk_policy_version"),
            AttendanceException.actual_checkout_time.label("actual_checkout_time"),
            AttendanceException.extended_deadline_at.label("extended_deadline_at"),
            AttendanceException.created_at.label("created_at"),
            AttendanceException.detected_at.label("detected_at"),
            AttendanceException.expires_at.label("expires_at"),
            AttendanceException.employee_explanation.label("employee_explanation"),
            AttendanceException.employee_submitted_at.label("employee_submitted_at"),
            AttendanceException.admin_note.label("admin_note"),
            AttendanceException.admin_decided_at.label("admin_decided_at"),
            AttendanceException.decided_by.label("decided_by"),
            decider.email.label("decided_by_email"),
            AttendanceException.resolved_at.label("resolved_at"),
            AttendanceException.resolved_by.label("resolved_by"),
            resolver.email.label("resolved_by_email"),
        )
        .join(Employee, Employee.id == AttendanceException.employee_id)
        .outerjoin(Group, Group.id == Employee.group_id)
        .outerjoin(AttendanceLog, AttendanceLog.id == AttendanceException.source_checkin_log_id)
        .outerjoin(decider, decider.id == AttendanceException.decided_by)
        .outerjoin(resolver, resolver.id == AttendanceException.resolved_by)
        .filter(AttendanceException.id == exception_id)
        .first()
    )

    if row is None:
        raise HTTPException(status_code=404, detail="attendance_exception not found")

    return AttendanceExceptionDetailResponse(
        id=row.id,
        employee_id=row.employee_id,
        employee_code=row.employee_code,
        full_name=row.full_name,
        group_code=row.group_code,
        group_name=row.group_name,
        work_date=row.work_date,
        exception_type=_normalize_exception_type(row.exception_type),
        status=normalize_exception_status(row.status),
        note=row.note,
        resolved_note=row.resolved_note,
        risk_score=row.risk_score,
        risk_level=row.risk_level,
        risk_flags=_deserialize_risk_flags(row.risk_flags),
        risk_policy_version=row.risk_policy_version,
        source_checkin_log_id=row.source_checkin_log_id,
        source_checkin_time=row.source_checkin_time,
        detected_at=row.detected_at,
        expires_at=row.expires_at,
        extended_deadline_at=row.extended_deadline_at,
        employee_explanation=row.employee_explanation,
        employee_submitted_at=row.employee_submitted_at,
        admin_note=row.admin_note,
        admin_decided_at=row.admin_decided_at,
        decided_by=row.decided_by,
        decided_by_email=row.decided_by_email,
        actual_checkout_time=row.actual_checkout_time,
        created_at=row.created_at,
        resolved_at=row.resolved_at,
        resolved_by=row.resolved_by,
        resolved_by_email=row.resolved_by_email,
        can_submit_explanation=_can_submit_explanation(row.status),
        can_admin_decide=_can_admin_decide(row.status),
        can_expire=_can_expire(row.status, row.expires_at),
        timeline=_build_exception_timeline(db, row.id),
    )


def _get_exception_or_404(db: Session, exception_id: int) -> AttendanceException:
    exception = db.query(AttendanceException).filter(AttendanceException.id == exception_id).first()
    if exception is None:
        raise HTTPException(status_code=404, detail="attendance_exception not found")
    return exception


def _get_employee_for_user(db: Session, user: User) -> Employee:
    employee = db.query(Employee).filter(Employee.user_id == user.id).first()
    if employee is None:
        raise HTTPException(status_code=400, detail="User is not linked to an employee")
    return employee


def _queue_employee_exception_notification(
    background_tasks: BackgroundTasks,
    db: Session,
    *,
    event_type: str,
    exception: AttendanceException,
    employee: Employee,
    admin_user: User | None = None,
    extra_metadata: dict[str, object] | None = None,
) -> None:
    if employee.user_id is None:
        return
    employee_user = db.query(User).filter(User.id == employee.user_id).first()
    if employee_user is None or not employee_user.email:
        return
    payload = build_exception_notification_mail(
        event_type=event_type,
        to_email=employee_user.email,
        exception=exception,
        employee=employee,
        recipient_role="EMPLOYEE",
        admin_user=admin_user,
        extra_metadata=extra_metadata,
    )
    if payload is None:
        return
    notification = create_exception_notification_record(
        db,
        payload=payload,
        exception_id=exception.id,
        recipient_user_id=employee_user.id,
        recipient_role="EMPLOYEE",
        dedupe_key=f"exception:{exception.id}:{event_type}:employee:{employee_user.id}",
    )
    if notification is not None:
        background_tasks.add_task(
            send_exception_notification_background,
            payload,
            notification.id,
            employee_user.fcm_token,
        )


def _queue_admin_exception_notifications(
    background_tasks: BackgroundTasks,
    db: Session,
    *,
    event_type: str,
    exception: AttendanceException,
    employee: Employee,
    extra_metadata: dict[str, object] | None = None,
) -> None:
    admins = db.query(User).filter(User.role == "ADMIN").all()
    for admin in admins:
        if not admin.email:
            continue
        admin_metadata = dict(extra_metadata or {})
        admin_metadata["employee_email"] = None
        payload = build_exception_notification_mail(
            event_type=event_type,
            to_email=admin.email,
            exception=exception,
            employee=employee,
            recipient_role="ADMIN",
            admin_user=admin,
            extra_metadata=admin_metadata,
        )
        if payload is None:
            continue
        notification = create_exception_notification_record(
            db,
            payload=payload,
            exception_id=exception.id,
            recipient_user_id=admin.id,
            recipient_role="ADMIN",
            dedupe_key=f"exception:{exception.id}:{event_type}:admin:{admin.id}",
        )
        if notification is not None:
            background_tasks.add_task(
                send_exception_notification_background,
                payload,
                notification.id,
                admin.fcm_token,
            )


def _normalize_action_note(value: str | None) -> str | None:
    if value is None:
        return None
    normalized = value.strip()
    return normalized or None


def _default_expires_at(detected_at: datetime, status: str, exception_type: str | None = None, db: Session | None = None) -> datetime | None:
    """Calculate expires_at for a new exception using the configured policy."""
    if status != PENDING_EMPLOYEE:
        return None
    hours = 72  # fallback if no DB / no policy
    if db is not None and exception_type is not None:
        policy = db.query(ExceptionPolicy).filter(ExceptionPolicy.id == 1).first()
        if policy is not None:
            hours = get_deadline_hours(policy, exception_type)
    return detected_at + timedelta(hours=hours)


def _expire_overdue_now(db: Session) -> int:
    """Bulk-expire all PENDING_EMPLOYEE exceptions whose effective deadline has passed.

    Effective deadline = extended_deadline_at if set, else expires_at.
    Returns number of records expired.
    """
    now = datetime.now(timezone.utc)
    count = (
        db.query(AttendanceException)
        .filter(
            AttendanceException.status == PENDING_EMPLOYEE,
            or_(
                and_(
                    AttendanceException.extended_deadline_at.isnot(None),
                    AttendanceException.extended_deadline_at < now,
                ),
                and_(
                    AttendanceException.extended_deadline_at.is_(None),
                    AttendanceException.expires_at.isnot(None),
                    AttendanceException.expires_at < now,
                ),
            ),
        )
        .update({"status": EXPIRED}, synchronize_session="evaluate")
    )
    if count > 0:
        db.flush()
    return count

def _to_excel_date(value: date | datetime | str | None) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        return value
    return value.isoformat()


def _to_excel_datetime(value: datetime | str | None) -> str | None:
    if value is None:
        return None

    if isinstance(value, str):
        try:
            parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
        except ValueError:
            return value
        value = parsed

    if value.tzinfo is None:
        value = value.replace(tzinfo=UTC_TZ)

    return value.astimezone(VN_TZ).strftime("%Y-%m-%d %H:%M:%S")


# ---------------------------------------------------------------------------
# Dashboard endpoints consumed by the Flutter admin panel
# ---------------------------------------------------------------------------


@router.get("/dashboard")
def get_dashboard_summary(
    date_param: date = Query(..., alias="date"),
    group_id: int | None = None,
    status: str | None = None,
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    """KPI summary cards for the admin dashboard."""

    # -- total employees (optionally filtered by group) --
    total_emp_q = db.query(func.count(Employee.id))
    if group_id:
        total_emp_q = total_emp_q.filter(Employee.group_id == group_id)
    total_employees: int = total_emp_q.scalar() or 0

    # -- checked-in employees on the given date --
    checked_in_q = (
        db.query(func.count(func.distinct(AttendanceLog.employee_id)))
        .filter(AttendanceLog.work_date == date_param, AttendanceLog.type == "IN")
    )
    if group_id:
        checked_in_q = checked_in_q.join(
            Employee, Employee.id == AttendanceLog.employee_id
        ).filter(Employee.group_id == group_id)
    checked_in: int = checked_in_q.scalar() or 0

    # -- late count --
    late_q = (
        db.query(func.count(AttendanceLog.id))
        .filter(
            AttendanceLog.work_date == date_param,
            AttendanceLog.type == "IN",
            AttendanceLog.punctuality_status == "LATE",
        )
    )
    if group_id:
        late_q = late_q.join(
            Employee, Employee.id == AttendanceLog.employee_id
        ).filter(Employee.group_id == group_id)
    late_count: int = late_q.scalar() or 0

    # -- out of range --
    oor_q = (
        db.query(func.count(AttendanceLog.id))
        .filter(
            AttendanceLog.work_date == date_param,
            AttendanceLog.type == "IN",
            AttendanceLog.is_out_of_range.is_(True),
        )
    )
    if group_id:
        oor_q = oor_q.join(
            Employee, Employee.id == AttendanceLog.employee_id
        ).filter(Employee.group_id == group_id)
    out_of_range_count: int = oor_q.scalar() or 0

    # -- geofence counts --
    active_geofence_q = db.query(func.count(GroupGeofence.id)).filter(GroupGeofence.active.is_(True))
    inactive_geofence_q = db.query(func.count(GroupGeofence.id)).filter(GroupGeofence.active.is_(False))
    geofence_count: int = active_geofence_q.scalar() or 0
    inactive_geofence_count: int = inactive_geofence_q.scalar() or 0

    attendance_rate = (checked_in / total_employees * 100) if total_employees > 0 else 0
    late_rate = (late_count / checked_in * 100) if checked_in > 0 else 0

    return {
        "total_employees": total_employees,
        "checked_in": checked_in,
        "attendance_rate": round(attendance_rate, 1),
        "late_count": late_count,
        "late_rate": round(late_rate, 1),
        "out_of_range_count": out_of_range_count,
        "geofence_count": geofence_count,
        "inactive_geofence_count": inactive_geofence_count,
        "employee_growth_percent": 0,
    }


@router.get("/attendance-logs")
def list_attendance_logs_for_dashboard(
    date_param: date | None = Query(None, alias="date"),
    from_date: date | None = Query(None, alias="from"),
    to_date: date | None = Query(None, alias="to"),
    group_id: int | None = None,
    status: str | None = None,
    search: str | None = None,
    sort: str | None = None,
    page: int | None = None,
    limit: int | None = None,
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    """Return attendance-log rows for the dashboard table."""

    effective_from = date_param or from_date
    effective_to = date_param or to_date

    rows = _fetch_daily_report_rows(db, effective_from, effective_to, None, group_id)
    exception_map = _build_exception_map(db, effective_from, effective_to, None, group_id)

    result: list[dict] = []
    for row in rows:
        checkin_status, checkout_status_val, attendance_state = _derive_daily_status(
            row.checkin_time, row.checkout_time,
            row.punctuality_rank, row.checkout_rank,
            row.checkout_raw_status,
        )
        exc_status, exc_type = exception_map.get((row.employee_id, row.work_date), (None, None))
        attendance_state = _apply_exception_to_attendance_state(attendance_state, exc_status, exc_type)

        # Apply status filter
        if status and status != "all":
            normalised_filter = status.upper()
            if normalised_filter == "ON_TIME" and checkin_status != "ON_TIME":
                continue
            elif normalised_filter == "LATE" and checkin_status != "LATE":
                continue
            elif normalised_filter == "EARLY" and checkin_status != "EARLY":
                continue
            elif normalised_filter == "ABSENT" and attendance_state != "ABSENT":
                continue
            elif normalised_filter == "OUT_OF_RANGE" and not row.out_of_range:
                continue

        # Apply search filter
        if search and search.strip():
            needle = search.strip().lower()
            haystack = f"{row.employee_code} {row.full_name} {row.group_name or ''}".lower()
            if needle not in haystack:
                continue

        checkin_vn = to_vn_time(row.checkin_time).strftime("%H:%M") if row.checkin_time else "--:--"
        checkout_vn = to_vn_time(row.checkout_time).strftime("%H:%M") if row.checkout_time else "--:--"

        total_hours = "--"
        if row.checkin_time and row.checkout_time:
            delta = row.checkout_time - row.checkin_time
            hours = max(0.0, delta.total_seconds() / 3600)
            total_hours = f"{hours:.1f}h"

        out_of_range_val = bool(row.out_of_range) if row.out_of_range is not None else False
        location_status = "outside" if out_of_range_val else "inside"

        # "status" is used by the dashboard badge (punctuality-based).
        # "attendance_status" keeps the full daily state.
        display_status = checkin_status or attendance_state or "on_time"

        result.append({
            "id": row.employee_id,
            "employee_name": row.full_name,
            "employee_code": row.employee_code,
            "department_name": row.group_name or "-",
            "group_name": row.group_name or "-",
            "work_date": row.work_date.isoformat() if row.work_date else None,
            "check_in_time": checkin_vn,
            "check_out_time": checkout_vn,
            "total_hours": total_hours,
            "location_status": location_status,
            "status": display_status,
            "attendance_status": attendance_state,
            "checkin_status": checkin_status,
            "checkout_status": checkout_status_val,
            "checkin_lat": float(row.checkin_lat) if row.checkin_lat is not None else None,
            "checkin_lng": float(row.checkin_lng) if row.checkin_lng is not None else None,
            "checkout_lat": float(row.checkout_lat) if row.checkout_lat is not None else None,
            "checkout_lng": float(row.checkout_lng) if row.checkout_lng is not None else None,
        })

    # Sort
    if sort:
        reverse = sort.startswith("-")
        sort_key = sort.lstrip("-")
        if sort_key in ("employee_name", "check_in_time", "attendance_status"):
            result.sort(key=lambda r: r.get(sort_key, ""), reverse=reverse)

    # Pagination
    total = len(result)
    if page and limit:
        start = (page - 1) * limit
        result = result[start : start + limit]

    return {"data": result, "total": total}


@router.get("/weekly-trends")
def get_weekly_trends(
    date_param: date = Query(..., alias="date"),
    group_id: int | None = None,
    status: str | None = None,
    period: str | None = None,
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    """Return on_time / late / out_of_range counts for the selected report period."""

    vn_day_names = {0: "T2", 1: "T3", 2: "T4", 3: "T5", 4: "T6", 5: "T7", 6: "CN"}

    normalised_period = (period or "day").strip().lower()
    if normalised_period not in {"day", "week", "month", "weekday"}:
        raise HTTPException(status_code=400, detail="period must be one of: day, week, month, weekday")

    normalised_status = (status or "all").strip().lower()
    if normalised_status not in {"", "all", "on_time", "late", "out_of_range"}:
        raise HTTPException(status_code=400, detail="status must be one of: all, on_time, late, out_of_range")

    def month_start(value: date) -> date:
        return date(value.year, value.month, 1)

    def month_end(value: date) -> date:
        if value.month == 12:
            return date(value.year, 12, 31)
        return date(value.year, value.month + 1, 1) - timedelta(days=1)

    def add_months(value: date, delta: int) -> date:
        month_index = value.year * 12 + (value.month - 1) + delta
        return date(month_index // 12, month_index % 12 + 1, 1)

    buckets: list[tuple[object, str, str]] = []
    bucket_by_date: dict[date, object] = {}

    if normalised_period == "month":
        selected_month = month_start(date_param)
        start_date = add_months(selected_month, -5)
        end_date = month_end(selected_month)
        current = start_date
        while current <= selected_month:
            label = current.strftime("%m/%Y")
            buckets.append((current, label, label))
            current = add_months(current, 1)
    elif normalised_period == "weekday":
        start_date = month_start(date_param)
        end_date = month_end(date_param)
        weekday_buckets = (
            (0, "T2"),
            (1, "T3"),
            (2, "T4"),
            (3, "T5"),
            (4, "T6"),
            (5, "T7"),
        )
        for weekday_index, label in weekday_buckets:
            buckets.append((weekday_index, label, label))
    else:
        start_date = month_start(date_param)
        end_date = month_end(date_param)
        if normalised_period == "day":
            current = start_date
            while current <= end_date:
                weekday_idx = current.weekday()
                day_label = vn_day_names.get(weekday_idx, current.strftime("%d/%m"))
                buckets.append((current, current.strftime("%d/%m"), f"{day_label} {current.strftime('%d/%m')}"))
                current += timedelta(days=1)
        else:
            current = start_date
            index = 1
            while current <= end_date:
                bucket_end = min(current + timedelta(days=6), end_date)
                day = f"T{index}"
                buckets.append((current, day, f"{day} {current.strftime('%d/%m')}-{bucket_end.strftime('%d/%m')}"))
                current = bucket_end + timedelta(days=1)
                index += 1

    if normalised_period == "month":
        current = start_date
        while current <= end_date:
            bucket_by_date[current] = month_start(current)
            current += timedelta(days=1)
    elif normalised_period == "weekday":
        current = start_date
        while current <= end_date:
            weekday_index = current.weekday()
            if 0 <= weekday_index <= 5:
                bucket_by_date[current] = weekday_index
            current += timedelta(days=1)
    elif normalised_period == "week":
        for bucket_start, _, _ in buckets:
            bucket_end = min(bucket_start + timedelta(days=6), end_date)
            current = bucket_start
            while current <= bucket_end:
                bucket_by_date[current] = bucket_start
                current += timedelta(days=1)
    else:
        for bucket_start, _, _ in buckets:
            bucket_by_date[bucket_start] = bucket_start

    work_date_expr = _work_date_expr(db)

    base_q = (
        db.query(
            work_date_expr.label("wd"),
            AttendanceLog.punctuality_status,
            AttendanceLog.is_out_of_range,
        )
        .join(Employee, Employee.id == AttendanceLog.employee_id)
        .filter(
            AttendanceLog.type == "IN",
            work_date_expr >= start_date,
            work_date_expr <= end_date,
        )
    )
    if group_id:
        base_q = base_q.filter(Employee.group_id == group_id)
    if normalised_status == "on_time":
        base_q = base_q.filter(AttendanceLog.punctuality_status.in_(("ON_TIME", "EARLY")))
    elif normalised_status == "late":
        base_q = base_q.filter(AttendanceLog.punctuality_status == "LATE")
    elif normalised_status == "out_of_range":
        base_q = base_q.filter(AttendanceLog.is_out_of_range.is_(True))

    rows = base_q.all()

    bucket_stats: dict[object, dict[str, int]] = {
        key: {"on_time": 0, "late": 0, "out_of_range": 0} for key, _, _ in buckets
    }

    for row in rows:
        wd = row.wd
        if isinstance(wd, str):
            wd = date.fromisoformat(wd)
        elif isinstance(wd, datetime):
            wd = wd.date()
        bucket_key = bucket_by_date.get(wd)
        if bucket_key is None:
            continue
        if row.punctuality_status == "ON_TIME" or row.punctuality_status == "EARLY":
            bucket_stats[bucket_key]["on_time"] += 1
        elif row.punctuality_status == "LATE":
            bucket_stats[bucket_key]["late"] += 1
        if row.is_out_of_range:
            bucket_stats[bucket_key]["out_of_range"] += 1

    result = []
    for bucket_key, day, day_label in buckets:
        stats = bucket_stats[bucket_key]
        result.append({
            "day": day,
            "day_label": day_label,
            "on_time": stats["on_time"],
            "on_time_count": stats["on_time"],
            "late": stats["late"],
            "out_of_range": stats["out_of_range"],
            "oor": stats["out_of_range"],
        })

    return result


@router.get("/exceptions")
def list_dashboard_exceptions(
    status_filter: str = Query("pending", alias="status"),
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    """Simplified exception list for the dashboard sidebar."""

    q = (
        db.query(
            AttendanceException.id,
            Employee.full_name,
            AttendanceException.exception_type,
            AttendanceException.status,
            AttendanceException.created_at,
        )
        .join(Employee, Employee.id == AttendanceException.employee_id)
    )

    normalised = status_filter.strip().upper() if status_filter else ""
    if normalised in {"PENDING", "OPEN"}:
        q = q.filter(AttendanceException.status.in_(["OPEN", PENDING_EMPLOYEE, PENDING_ADMIN]))
    elif normalised == "RESOLVED":
        q = q.filter(AttendanceException.status.in_(["RESOLVED", APPROVED, REJECTED, EXPIRED]))
    elif normalised in {PENDING_EMPLOYEE, PENDING_ADMIN, APPROVED, REJECTED, EXPIRED}:
        q = q.filter(AttendanceException.status == normalised)

    rows = q.order_by(AttendanceException.created_at.desc()).limit(20).all()

    return [
        {
            "id": row.id,
            "name": row.full_name,
            "employee_name": row.full_name,
            "full_name": row.full_name,
            "reason": row.exception_type,
            "exception_type": row.exception_type,
            "status": normalize_exception_status(row.status).lower(),
            "time": row.created_at.isoformat() if row.created_at else None,
            "created_at": row.created_at.isoformat() if row.created_at else None,
        }
        for row in rows
    ]


# ---------------------------------------------------------------------------


from pydantic import BaseModel as _BaseModel, Field as _Field


class _ExportExcelBody(_BaseModel):
    from_date: date | None = _Field(None, alias="from")
    to_date: date | None = _Field(None, alias="to")
    group_id: int | None = None
    status: str | None = None

    model_config = {"populate_by_name": True}


@router.post("/export-excel")
def export_excel_via_post(
    body: _ExportExcelBody,
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    """POST variant used by the Flutter dashboard CSV/Excel export button."""
    return export_attendance_report_excel(
        from_date=body.from_date,
        to_date=body.to_date,
        group_id=body.group_id,
        status=body.status,
        include_empty=True,
        db=db,
    )


@router.get("/attendance.xlsx")
def export_attendance_report_excel(
    from_date: date | None = Query(None, alias="from"),
    to_date: date | None = Query(None, alias="to"),
    employee_id: int | None = None,
    group_id: int | None = None,
    status: str | None = None,
    search: str | None = None,
    include_empty: bool = False,
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    if from_date and to_date and from_date > to_date:
        raise HTTPException(status_code=400, detail="'from' must be <= 'to'")

    if employee_id is not None:
        emp = db.query(Employee).filter(Employee.id == employee_id).first()
        if not emp:
            raise HTTPException(status_code=404, detail="employee_id not found")

    if group_id is not None:
        group = db.query(Group).filter(Group.id == group_id).first()
        if not group:
            raise HTTPException(status_code=404, detail="group_id not found")

    rows = _fetch_daily_report_rows(db, from_date, to_date, employee_id, group_id)
    exception_status_map = _build_exception_map(db, from_date, to_date, employee_id, group_id)
    if status and status != "all":
        normalised_filter = status.upper()
        filtered_rows = []
        for row in rows:
            checkin_status, _checkout_status, attendance_state = _derive_daily_status(
                row.checkin_time,
                row.checkout_time,
                row.punctuality_rank,
                row.checkout_rank,
                row.checkout_raw_status,
            )
            exc_status, exc_type = exception_status_map.get((row.employee_id, row.work_date), (None, None))
            attendance_state = _apply_exception_to_attendance_state(attendance_state, exc_status, exc_type)
            if normalised_filter == "ON_TIME" and checkin_status != "ON_TIME":
                continue
            if normalised_filter == "LATE" and checkin_status != "LATE":
                continue
            if normalised_filter == "EARLY" and checkin_status != "EARLY":
                continue
            if normalised_filter == "ABSENT" and attendance_state != "ABSENT":
                continue
            if normalised_filter == "OUT_OF_RANGE" and not row.out_of_range:
                continue
            filtered_rows.append(row)
        rows = filtered_rows

    if search and search.strip():
        needle = search.strip().lower()
        rows = [
            row
            for row in rows
            if needle in f"{row.employee_code} {row.full_name} {row.group_name or ''}".lower()
        ]

    if not rows and not include_empty:
        raise HTTPException(status_code=404, detail="No attendance data for selected filters")

    active_rule = db.query(CheckinRule).filter(CheckinRule.active.is_(True)).first()
    fallback_radius_m = active_rule.radius_m if active_rule is not None else None
    group_ids = {int(row.group_id) for row in rows if row.group_id is not None}
    geofence_radius_map, group_max_radius_map = load_group_geofence_radius_maps(db, group_ids)

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    headers = [
        "date",
        "employee_code",
        "full_name",
        "group_code",
        "group_name",
        "matched_geofence",
        "geofence_source",
        "checkin_time",
        "checkout_time",
        "checkin_status",
        "checkout_status",
        "attendance_state",
        "out_of_range",
        "avg_distance_m",
        "max_distance_m",
        "radius_m",
        "distance_consistency_warning",
        "regular_minutes",
        "overtime_minutes",
        "payable_overtime_minutes",
        "overtime_cross_day",
        "exception_status",
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    fill_ok = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    fill_warn = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    fill_audit = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

    for row in rows:
        out_of_range_value = bool(row.out_of_range) if row.out_of_range is not None else False
        range_status_text = "OUT_OF_RANGE" if out_of_range_value else "IN_RANGE"
        checkin_status, checkout_status, attendance_state = _derive_daily_status(
            row.checkin_time,
            row.checkout_time,
            row.punctuality_rank,
            row.checkout_rank,
            row.checkout_raw_status,
        )
        matched_geofence = row.checkin_matched_geofence or row.checkout_matched_geofence
        geofence_source = _rank_to_geofence_source(row.geofence_source_rank)
        avg_distance_m = float(row.avg_distance_m) if row.avg_distance_m is not None else None
        max_distance_m = float(row.max_distance_m) if row.max_distance_m is not None else None
        radius_m = resolve_reference_radius_m(
            geofence_source=geofence_source,
            matched_geofence=matched_geofence,
            group_id=row.group_id,
            fallback_radius_m=fallback_radius_m,
            named_radius_map=geofence_radius_map,
            max_radius_map=group_max_radius_map,
        )
        distance_consistency_warning = compute_distance_consistency_warning(
            out_of_range=out_of_range_value,
            avg_distance_m=avg_distance_m,
            max_distance_m=max_distance_m,
            radius_m=radius_m,
        )

        shift_start = row.shift_start or time(8, 0)
        shift_end = row.shift_end or time(17, 0)
        regular_minutes, overtime_minutes, overtime_cross_day = split_regular_overtime_minutes(
            row.work_date,
            row.checkin_time,
            row.checkout_time,
            shift_start,
            shift_end,
        )
        exception_status, exception_type = exception_status_map.get((row.employee_id, row.work_date), (None, None))
        attendance_state = _apply_exception_to_attendance_state(
            attendance_state=attendance_state,
            exception_status=exception_status,
            exception_type=exception_type,
        )
        payable_overtime_minutes = _compute_payable_overtime_minutes(
            overtime_minutes=overtime_minutes,
            exception_status=exception_status,
            exception_type=exception_type,
        )

        ws.append(
            [
                _to_excel_date(row.work_date),
                row.employee_code,
                row.full_name,
                row.group_code,
                row.group_name,
                matched_geofence,
                geofence_source,
                _to_excel_datetime(row.checkin_time),
                _to_excel_datetime(row.checkout_time),
                checkin_status,
                checkout_status,
                attendance_state,
                range_status_text,
                avg_distance_m,
                max_distance_m,
                radius_m,
                distance_consistency_warning,
                regular_minutes,
                overtime_minutes,
                payable_overtime_minutes,
                "YES" if overtime_cross_day else "NO",
                exception_status,
            ]
        )

        current_row_idx = ws.max_row
        out_of_range_col = headers.index("out_of_range") + 1
        range_cell = ws.cell(row=current_row_idx, column=out_of_range_col)
        range_cell.fill = fill_warn if out_of_range_value else fill_ok

        if distance_consistency_warning:
            warning_col = headers.index("distance_consistency_warning") + 1
            ws.cell(row=current_row_idx, column=warning_col).fill = fill_audit

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    for col_idx, _ in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = "attendance_report.xlsx"
    if group_id is not None:
        filename = f"attendance_report_group_{group_id}.xlsx"
    if from_date or to_date:
        filename = (
            f"attendance_report_{from_date or 'all'}_{to_date or 'all'}"
            + (f"_group_{group_id}" if group_id is not None else "")
            + ".xlsx"
        )

    headers_response = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_response,
    )


@router.get("/attendance-exceptions", response_model=list[AttendanceExceptionReportResponse])
def list_attendance_exceptions(
    from_date: date | None = Query(None, alias="from"),
    to_date: date | None = Query(None, alias="to"),
    employee_id: int | None = None,
    group_id: int | None = None,
    exception_type: str = "MISSED_CHECKOUT",
    status_filter: str | None = Query(None, alias="status"),
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    if from_date and to_date and from_date > to_date:
        raise HTTPException(status_code=400, detail="'from' must be <= 'to'")

    normalized_exception_type = exception_type.strip().upper()
    if normalized_exception_type == "GPS_RISK":
        normalized_exception_type = "SUSPECTED_LOCATION_SPOOF"
    if normalized_exception_type not in {"MISSED_CHECKOUT", "AUTO_CLOSED", "SUSPECTED_LOCATION_SPOOF", "LARGE_TIME_DEVIATION"}:
        raise HTTPException(
            status_code=400,
            detail="exception_type must be MISSED_CHECKOUT, AUTO_CLOSED, SUSPECTED_LOCATION_SPOOF or LARGE_TIME_DEVIATION",
        )

    normalized_status = status_filter.strip().upper() if status_filter else None
    if normalized_status == "OPEN":
        normalized_statuses = [PENDING_EMPLOYEE]
    elif normalized_status == "RESOLVED":
        normalized_statuses = [APPROVED]
    elif normalized_status == "PENDING":
        normalized_statuses = [PENDING_EMPLOYEE, PENDING_ADMIN]
    elif normalized_status:
        try:
            normalized_statuses = build_exception_status_filter_values([normalized_status])
        except ValueError as exc_error:
            raise HTTPException(status_code=400, detail=str(exc_error))
    else:
        normalized_statuses = None

    # Lazy expiry: flip any overdue PENDING_EMPLOYEE to EXPIRED before querying
    _expire_overdue_now(db)

    resolver = aliased(User)
    decider = aliased(User)

    q = (
        db.query(
            AttendanceException.id.label("id"),
            AttendanceException.employee_id.label("employee_id"),
            Employee.code.label("employee_code"),
            Employee.full_name.label("full_name"),
            Group.id.label("group_id"),
            Group.code.label("group_code"),
            Group.name.label("group_name"),
            AttendanceException.work_date.label("work_date"),
            AttendanceException.exception_type.label("exception_type"),
            AttendanceException.status.label("status"),
            AttendanceException.note.label("note"),
            AttendanceException.resolved_note.label("resolved_note"),
            AttendanceException.source_checkin_log_id.label("source_checkin_log_id"),
            AttendanceLog.time.label("source_checkin_time"),
            AttendanceLog.risk_score.label("risk_score"),
            AttendanceLog.risk_level.label("risk_level"),
            AttendanceLog.risk_flags.label("risk_flags"),
            AttendanceLog.risk_policy_version.label("risk_policy_version"),
            AttendanceException.actual_checkout_time.label("actual_checkout_time"),
            AttendanceException.extended_deadline_at.label("extended_deadline_at"),
            AttendanceException.created_at.label("created_at"),
            AttendanceException.detected_at.label("detected_at"),
            AttendanceException.expires_at.label("expires_at"),
            AttendanceException.employee_explanation.label("employee_explanation"),
            AttendanceException.employee_submitted_at.label("employee_submitted_at"),
            AttendanceException.admin_note.label("admin_note"),
            AttendanceException.admin_decided_at.label("admin_decided_at"),
            AttendanceException.decided_by.label("decided_by"),
            decider.email.label("decided_by_email"),
            AttendanceException.resolved_at.label("resolved_at"),
            AttendanceException.resolved_by.label("resolved_by"),
            resolver.email.label("resolved_by_email"),
        )
        .join(Employee, Employee.id == AttendanceException.employee_id)
        .outerjoin(Group, Group.id == Employee.group_id)
        .outerjoin(AttendanceLog, AttendanceLog.id == AttendanceException.source_checkin_log_id)
        .outerjoin(decider, decider.id == AttendanceException.decided_by)
        .outerjoin(resolver, resolver.id == AttendanceException.resolved_by)
    )
    if normalized_exception_type == "SUSPECTED_LOCATION_SPOOF":
        q = q.filter(AttendanceException.exception_type.in_(["SUSPECTED_LOCATION_SPOOF", "GPS_RISK"]))
    else:
        q = q.filter(AttendanceException.exception_type == normalized_exception_type)

    if employee_id:
        q = q.filter(AttendanceException.employee_id == employee_id)
    if group_id:
        q = q.filter(Employee.group_id == group_id)
    if from_date:
        q = q.filter(AttendanceException.work_date >= from_date)
    if to_date:
        q = q.filter(AttendanceException.work_date <= to_date)
    if normalized_statuses:
        status_options = list(normalized_statuses)
        if PENDING_EMPLOYEE in normalized_statuses:
            status_options.append("OPEN")
        if APPROVED in normalized_statuses:
            status_options.append("RESOLVED")
        q = q.filter(AttendanceException.status.in_(status_options))

    rows = q.order_by(AttendanceException.work_date.desc(), AttendanceException.created_at.desc()).all()

    return [
        AttendanceExceptionReportResponse(
            id=row.id,
            employee_id=row.employee_id,
            employee_code=row.employee_code,
            full_name=row.full_name,
            group_code=row.group_code,
            group_name=row.group_name,
            work_date=row.work_date,
            exception_type=_normalize_exception_type(row.exception_type),
            status=normalize_exception_status(row.status),
            note=row.note,
            resolved_note=row.resolved_note,
            risk_score=row.risk_score,
            risk_level=row.risk_level,
            risk_flags=_deserialize_risk_flags(row.risk_flags),
            risk_policy_version=row.risk_policy_version,
            source_checkin_log_id=row.source_checkin_log_id,
            source_checkin_time=row.source_checkin_time,
            detected_at=row.detected_at,
            expires_at=row.expires_at,
            extended_deadline_at=row.extended_deadline_at,
            employee_explanation=row.employee_explanation,
            employee_submitted_at=row.employee_submitted_at,
            admin_note=row.admin_note,
            admin_decided_at=row.admin_decided_at,
            decided_by=row.decided_by,
            decided_by_email=row.decided_by_email,
            actual_checkout_time=row.actual_checkout_time,
            created_at=row.created_at,
            resolved_at=row.resolved_at,
            resolved_by=row.resolved_by,
            resolved_by_email=row.resolved_by_email,
            can_submit_explanation=_can_submit_explanation(row.status),
            can_admin_decide=_can_admin_decide(row.status),
            can_expire=_can_expire(row.status, row.expires_at),
        )
        for row in rows
    ]


@router.get("/attendance-exceptions/me", response_model=list[AttendanceExceptionReportResponse])
def list_my_attendance_exceptions(
    status_filter: str | None = Query(None, alias="status"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    employee = _get_employee_for_user(db, current_user)

    normalized_status = status_filter.strip().upper() if status_filter else None
    if normalized_status == "OPEN":
        normalized_statuses = [PENDING_EMPLOYEE]
    elif normalized_status == "RESOLVED":
        normalized_statuses = [APPROVED]
    elif normalized_status == "PENDING":
        normalized_statuses = [PENDING_EMPLOYEE, PENDING_ADMIN]
    elif normalized_status:
        try:
            normalized_statuses = build_exception_status_filter_values([normalized_status])
        except ValueError as exc_error:
            raise HTTPException(status_code=400, detail=str(exc_error))
    else:
        normalized_statuses = None

    # Lazy expiry: flip any overdue PENDING_EMPLOYEE to EXPIRED before querying
    _expire_overdue_now(db)

    q = db.query(AttendanceException.id).filter(AttendanceException.employee_id == employee.id)
    if normalized_statuses:
        status_options = list(normalized_statuses)
        if PENDING_EMPLOYEE in normalized_statuses:
            status_options.append("OPEN")
        if APPROVED in normalized_statuses:
            status_options.append("RESOLVED")
        q = q.filter(AttendanceException.status.in_(status_options))

    rows = q.order_by(AttendanceException.work_date.desc(), AttendanceException.created_at.desc()).all()
    return [_build_exception_response(db, row.id) for row in rows]


@router.get("/attendance-exceptions/me/{exception_id}", response_model=AttendanceExceptionDetailResponse)
def get_my_attendance_exception_detail(
    exception_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    employee = _get_employee_for_user(db, current_user)
    exception = _get_exception_or_404(db, exception_id)
    if exception.employee_id != employee.id:
        raise HTTPException(status_code=403, detail="Employee can only view owned exception")
    auto_expire_overdue(db, [exception])
    return _build_exception_response(db, exception.id)


@router.get("/attendance-exceptions/{exception_id}", response_model=AttendanceExceptionDetailResponse)
def get_attendance_exception_detail(
    exception_id: int,
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    exception = _get_exception_or_404(db, exception_id)
    auto_expire_overdue(db, [exception])
    return _build_exception_response(db, exception.id)


@router.post("/attendance-exceptions/system", response_model=AttendanceExceptionDetailResponse)
def create_attendance_exception(
    payload: AttendanceExceptionCreateRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    _system_actor: str = Depends(require_exception_workflow_system),
):
    source_checkin = db.query(AttendanceLog).filter(AttendanceLog.id == payload.source_checkin_log_id).first()
    if source_checkin is None:
        raise HTTPException(status_code=400, detail="source_checkin_log_id not found")
    if source_checkin.employee_id != payload.employee_id:
        raise HTTPException(status_code=400, detail="source_checkin_log_id does not belong to employee_id")

    employee = db.query(Employee).filter(Employee.id == payload.employee_id).first()
    if employee is None:
        raise HTTPException(status_code=404, detail="employee_id not found")

    existing = (
        db.query(AttendanceException)
        .filter(AttendanceException.source_checkin_log_id == payload.source_checkin_log_id)
        .first()
    )
    if existing is not None:
        raise HTTPException(status_code=409, detail="attendance_exception already exists for source_checkin_log_id")

    detected_at = normalize_utc(payload.detected_at) if payload.detected_at is not None else datetime.now(timezone.utc)
    initial_status = default_exception_status_for_type(payload.exception_type)
    expires_at = normalize_utc(payload.expires_at) if payload.expires_at is not None else _default_expires_at(detected_at, initial_status, payload.exception_type, db)
    if expires_at is not None and expires_at <= detected_at:
        raise HTTPException(status_code=400, detail="expires_at must be later than detected_at")

    exception = AttendanceException(
        employee_id=payload.employee_id,
        source_checkin_log_id=payload.source_checkin_log_id,
        exception_type=payload.exception_type,
        work_date=payload.work_date or source_checkin.work_date or source_checkin.time.date(),
        status=initial_status,
        note=_normalize_action_note(payload.note),
        detected_at=detected_at,
        expires_at=expires_at,
    )
    db.add(exception)
    db.flush()
    record_attendance_exception_audit(
        db,
        exception_id=exception.id,
        event_type="exception_detected",
        previous_status=None,
        next_status=exception.status,
        actor_type="SYSTEM",
        actor_email="SYSTEM",
        metadata={
            "exception_type": exception.exception_type,
            "source_checkin_log_id": exception.source_checkin_log_id,
            "employee_id": exception.employee_id,
        },
    )
    if exception.status == PENDING_EMPLOYEE:
        _queue_employee_exception_notification(
            background_tasks,
            db,
            event_type="exception_detected_employee",
            exception=exception,
            employee=employee,
        )
    elif exception.status == PENDING_ADMIN:
        _queue_admin_exception_notifications(
            background_tasks,
            db,
            event_type="exception_detected_admin",
            exception=exception,
            employee=employee,
        )
    db.commit()
    db.refresh(exception)
    return _build_exception_response(db, exception.id)


@router.post("/attendance-exceptions/{exception_id}/submit-explanation", response_model=AttendanceExceptionDetailResponse)
def submit_attendance_exception_explanation(
    exception_id: int,
    payload: AttendanceExceptionSubmitExplanationRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    employee = _get_employee_for_user(db, current_user)
    exception = _get_exception_or_404(db, exception_id)
    if exception.employee_id != employee.id:
        raise HTTPException(status_code=403, detail="Employee can only submit explanation for owned exception")

    explanation = payload.explanation.strip()
    if not explanation:
        raise HTTPException(status_code=400, detail="explanation must not be empty")

    # Block submission after effective deadline
    effective_deadline = get_effective_deadline(exception)
    if effective_deadline is not None:
        now_utc = datetime.now(timezone.utc)
        deadline_utc = effective_deadline if effective_deadline.tzinfo else effective_deadline.replace(tzinfo=timezone.utc)
        if now_utc > deadline_utc:
            exception.status = EXPIRED
            db.flush()
            raise HTTPException(status_code=410, detail="Đã quá hạn giải trình. Vui lòng liên hệ quản trị viên để được gia hạn.")

    previous_status = exception.status
    try:
        exception.status = ensure_allowed_exception_transition(exception.status, PENDING_ADMIN)
    except ValueError as exc_error:
        raise HTTPException(status_code=409, detail=str(exc_error))

    submitted_at = datetime.now(timezone.utc)
    exception.employee_explanation = explanation
    exception.employee_submitted_at = submitted_at
    record_attendance_exception_audit(
        db,
        exception_id=exception.id,
        event_type="employee_explanation_submitted",
        previous_status=previous_status,
        next_status=exception.status,
        actor_type="EMPLOYEE",
        actor_id=current_user.id,
        actor_email=current_user.email,
        metadata={
            "employee_id": employee.id,
            "employee_submitted_at": submitted_at.isoformat(),
        },
    )
    _queue_admin_exception_notifications(
        background_tasks,
        db,
        event_type="exception_submitted_admin",
        exception=exception,
        employee=employee,
        extra_metadata={
            "employee_explanation": explanation,
            "employee_submitted_at": submitted_at.isoformat(),
        },
    )
    db.commit()
    return _build_exception_response(db, exception.id)


@router.post("/attendance-exceptions/{exception_id}/approve", response_model=AttendanceExceptionDetailResponse)
def approve_attendance_exception(
    exception_id: int,
    payload: AttendanceExceptionApproveRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    admin_user: User = Depends(require_admin),
):
    exception = _get_exception_or_404(db, exception_id)
    source_checkin = db.query(AttendanceLog).filter(AttendanceLog.id == exception.source_checkin_log_id).first()
    if source_checkin is None:
        raise HTTPException(status_code=400, detail="source check-in log not found")

    normalized_exception_type = _normalize_exception_type(exception.exception_type)
    if normalized_exception_type in {"AUTO_CLOSED", "MISSED_CHECKOUT"} and payload.actual_checkout_time is not None:
        actual_checkout_utc = normalize_utc(payload.actual_checkout_time)
        if actual_checkout_utc <= normalize_utc(source_checkin.time):
            raise HTTPException(status_code=400, detail="actual_checkout_time must be later than source check-in time")
        _upsert_checkout_log_from_resolution(db, exception, source_checkin, actual_checkout_utc)
        exception.actual_checkout_time = actual_checkout_utc

    previous_status = exception.status
    try:
        exception.status = ensure_allowed_exception_transition(exception.status, APPROVED)
    except ValueError as exc_error:
        raise HTTPException(status_code=409, detail=str(exc_error))

    decision_time = datetime.now(timezone.utc)
    note = _normalize_action_note(payload.admin_note)
    exception.admin_note = note
    exception.admin_decided_at = decision_time
    exception.decided_by = admin_user.id
    exception.resolved_by = admin_user.id
    exception.resolved_at = decision_time
    exception.resolved_note = note
    record_attendance_exception_audit(
        db,
        exception_id=exception.id,
        event_type="admin_approved",
        previous_status=previous_status,
        next_status=exception.status,
        actor_type="ADMIN",
        actor_id=admin_user.id,
        actor_email=admin_user.email,
        metadata={
            "admin_decided_at": decision_time.isoformat(),
            "actual_checkout_time": exception.actual_checkout_time.isoformat() if exception.actual_checkout_time else None,
        },
    )
    employee = db.query(Employee).filter(Employee.id == exception.employee_id).first()
    if employee is not None:
        _queue_employee_exception_notification(
            background_tasks,
            db,
            event_type="exception_approved_employee",
            exception=exception,
            employee=employee,
            admin_user=admin_user,
        )
    db.commit()
    return _build_exception_response(db, exception.id)


@router.post("/attendance-exceptions/{exception_id}/reject", response_model=AttendanceExceptionDetailResponse)
def reject_attendance_exception(
    exception_id: int,
    payload: AttendanceExceptionRejectRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    admin_user: User = Depends(require_admin),
):
    exception = _get_exception_or_404(db, exception_id)
    admin_note = payload.admin_note.strip()
    if not admin_note:
        raise HTTPException(status_code=400, detail="admin_note is required")

    previous_status = exception.status
    try:
        exception.status = ensure_allowed_exception_transition(exception.status, REJECTED)
    except ValueError as exc_error:
        raise HTTPException(status_code=409, detail=str(exc_error))

    decision_time = datetime.now(timezone.utc)
    exception.admin_note = admin_note
    exception.admin_decided_at = decision_time
    exception.decided_by = admin_user.id
    exception.resolved_by = admin_user.id
    exception.resolved_at = decision_time
    exception.resolved_note = admin_note
    record_attendance_exception_audit(
        db,
        exception_id=exception.id,
        event_type="admin_rejected",
        previous_status=previous_status,
        next_status=exception.status,
        actor_type="ADMIN",
        actor_id=admin_user.id,
        actor_email=admin_user.email,
        metadata={
            "admin_decided_at": decision_time.isoformat(),
        },
    )
    employee = db.query(Employee).filter(Employee.id == exception.employee_id).first()
    if employee is not None:
        _queue_employee_exception_notification(
            background_tasks,
            db,
            event_type="exception_rejected_employee",
            exception=exception,
            employee=employee,
            admin_user=admin_user,
        )
    db.commit()
    return _build_exception_response(db, exception.id)


@router.post("/attendance-exceptions/{exception_id}/expire", response_model=AttendanceExceptionDetailResponse)
def expire_attendance_exception(
    exception_id: int,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    _system_actor: str = Depends(require_exception_workflow_system),
):
    exception = _get_exception_or_404(db, exception_id)
    if exception.expires_at is None:
        raise HTTPException(status_code=400, detail="expires_at is required to expire exception")

    now_utc = datetime.now(timezone.utc)
    if normalize_utc(exception.expires_at) > now_utc:
        raise HTTPException(status_code=400, detail="Exception cannot be expired before expires_at")

    previous_status = exception.status
    try:
        exception.status = ensure_allowed_exception_transition(exception.status, EXPIRED)
    except ValueError as exc_error:
        raise HTTPException(status_code=409, detail=str(exc_error))

    record_attendance_exception_audit(
        db,
        exception_id=exception.id,
        event_type="system_expired",
        previous_status=previous_status,
        next_status=exception.status,
        actor_type="SYSTEM",
        actor_email="SYSTEM",
        metadata={
            "expired_at": now_utc.isoformat(),
            "expires_at": normalize_utc(exception.expires_at).isoformat(),
        },
    )
    employee = db.query(Employee).filter(Employee.id == exception.employee_id).first()
    if employee is not None:
        _queue_employee_exception_notification(
            background_tasks,
            db,
            event_type="exception_expired_employee",
            exception=exception,
            employee=employee,
            extra_metadata={
                "expired_at": now_utc.isoformat(),
            },
        )
    db.commit()
    return _build_exception_response(db, exception.id)


@router.patch("/attendance-exceptions/{exception_id}/resolve", response_model=AttendanceExceptionDetailResponse)
def resolve_attendance_exception(
    exception_id: int,
    payload: AttendanceExceptionResolveRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    admin_user: User = Depends(require_admin),
):
    return approve_attendance_exception(
        exception_id=exception_id,
        payload=AttendanceExceptionApproveRequest(
            admin_note=payload.note,
            actual_checkout_time=payload.actual_checkout_time,
        ),
        background_tasks=background_tasks,
        db=db,
        admin_user=admin_user,
    )


@router.patch("/attendance-exceptions/{exception_id}/reopen", response_model=AttendanceExceptionDetailResponse)
def reopen_attendance_exception(
    exception_id: int,
    db: Session = Depends(get_db),
    _admin_user: User = Depends(require_admin),
):
    raise HTTPException(
        status_code=409,
        detail="Legacy reopen flow is disabled in the new exception workflow",
    )


@router.patch("/attendance-exceptions/{exception_id}/extend-deadline", response_model=AttendanceExceptionDetailResponse)
def extend_exception_deadline(
    exception_id: int,
    payload: AttendanceExceptionExtendDeadlineRequest,
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    """Extend the explanation deadline for a single exception.

    Allowed for PENDING_EMPLOYEE and EXPIRED exceptions.
    If exception is currently EXPIRED, it is revived back to PENDING_EMPLOYEE.
    The new deadline is: effective_deadline + extend_hours (forward from current deadline).
    """
    exception = _get_exception_or_404(db, exception_id)

    if exception.status not in (PENDING_EMPLOYEE, EXPIRED):
        raise HTTPException(
            status_code=409,
            detail=f"Cannot extend deadline for exception in status {exception.status}. Only PENDING_EMPLOYEE or EXPIRED allowed.",
        )

    # Compute the current effective deadline (starting point for extension)
    current_deadline = get_effective_deadline(exception)
    now_utc = datetime.now(timezone.utc)

    if current_deadline is None:
        # No deadline set — start from now
        base = now_utc
    else:
        deadline_utc = current_deadline if current_deadline.tzinfo else current_deadline.replace(tzinfo=timezone.utc)
        # If deadline is still in the future, extend from there; otherwise extend from now
        base = max(deadline_utc, now_utc)

    new_deadline = base + timedelta(hours=payload.extend_hours)
    exception.extended_deadline_at = new_deadline

    # Revive EXPIRED exceptions back to PENDING_EMPLOYEE
    if exception.status == EXPIRED:
        exception.status = PENDING_EMPLOYEE

    db.flush()
    db.commit()
    return _build_exception_response(db, exception.id)


@router.post("/attendance-exceptions/batch-expire", response_model=dict)
def batch_expire_attendance_exceptions(
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    """Manually trigger lazy expiry for all overdue PENDING_EMPLOYEE exceptions.

    Useful for admin dashboards or scheduled reconciliation.
    Returns { "expired_count": N }.
    """
    count = _expire_overdue_now(db)
    db.commit()
    return {"expired_count": count}


@router.post("/attendance-exceptions/purge-expired", response_model=dict)
def purge_expired_attendance_exceptions(
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    """Delete EXPIRED exceptions whose retention grace period has passed."""
    expired_count = _expire_overdue_now(db)
    policy = db.query(ExceptionPolicy).filter(ExceptionPolicy.id == 1).first()
    grace_period_days = policy.grace_period_days if policy is not None else 30
    cutoff = datetime.now(timezone.utc) - timedelta(days=grace_period_days)

    rows = (
        db.query(AttendanceException.id)
        .filter(
            AttendanceException.status == EXPIRED,
            or_(
                and_(
                    AttendanceException.extended_deadline_at.isnot(None),
                    AttendanceException.extended_deadline_at < cutoff,
                ),
                and_(
                    AttendanceException.extended_deadline_at.is_(None),
                    AttendanceException.expires_at.isnot(None),
                    AttendanceException.expires_at < cutoff,
                ),
            ),
        )
        .all()
    )
    ids = [row.id for row in rows]
    if not ids:
        db.commit()
        return {
            "deleted_count": 0,
            "expired_count": expired_count,
            "grace_period_days": grace_period_days,
        }

    db.query(AttendanceExceptionNotification).filter(
        AttendanceExceptionNotification.exception_id.in_(ids)
    ).delete(synchronize_session=False)
    db.query(AttendanceExceptionAudit).filter(
        AttendanceExceptionAudit.exception_id.in_(ids)
    ).delete(synchronize_session=False)
    db.query(AttendanceException).filter(
        AttendanceException.id.in_(ids)
    ).delete(synchronize_session=False)
    db.commit()
    return {
        "deleted_count": len(ids),
        "expired_count": expired_count,
        "grace_period_days": grace_period_days,
    }


# ---------------------------------------------------------------------------
# Monthly Attendance Matrix — Phase 1: data / logic helpers
# ---------------------------------------------------------------------------

def _geofence_type(name: str | None, location_type: str | None = None) -> str:
    """Classify a geofence into VP / SITE / SYSTEM_RULE.

    Prefer explicit location_type from DB; fall back to keyword-matching on name
    for legacy records created before the location_type column existed.
    """
    if location_type in ("VP", "SITE"):
        return location_type
    if not name:
        return "SYSTEM_RULE"
    # legacy fallback: guess from name keywords
    n = name.lower()
    if any(k in n for k in ["vp", "văn phòng", "van phong", "office", "hq"]):
        return "VP"
    return "SITE"


_FULL_CODES = {"V", "S", "X"}
_HALF_CODES = {"1/2V", "1/2S", "1/2T"}


def _build_leave_map(
    db: Session,
    from_date: date,
    to_date: date,
    group_id: int | None = None,
) -> dict[tuple[int, date], str]:
    """Return {(employee_id, day): code} for all APPROVED leave requests that
    overlap [from_date, to_date].

    code is 'P' for PAID and 'K' for UNPAID (treated same as absent but
    annotated so the caller can distinguish if needed).
    """
    q = (
        db.query(LeaveRequest.employee_id, LeaveRequest.start_date, LeaveRequest.end_date, LeaveRequest.leave_type)
        .filter(
            LeaveRequest.status == "APPROVED",
            LeaveRequest.start_date <= to_date,
            LeaveRequest.end_date >= from_date,
        )
    )
    if group_id is not None:
        q = q.join(Employee, Employee.id == LeaveRequest.employee_id).filter(
            Employee.group_id == group_id
        )

    leave_map: dict[tuple[int, date], str] = {}
    for eid, start, end, ltype in q.all():
        code = "P" if ltype == "PAID" else "K"
        cur = max(start, from_date)
        while cur <= min(end, to_date):
            leave_map[(int(eid), cur)] = code
            cur += timedelta(days=1)
    return leave_map


def _derive_cell_code(row, exception_map: dict, holiday_set: set, geofence_location_map: dict | None = None) -> str:
    """
    Derive the attendance cell symbol for one (employee, work_date) pair.
    Only called when row is not None (has checkin).

    Priority (highest → lowest):
    1. Public holiday + has checkin → V / S / X (treat as normal worked day)
    2. MISSED_CHECKOUT exception → 1/2T
    3. Checkin LATE → 1/2T
    4-8. Full / half day by geofence type
    """
    if row is None:
        # Kept for safety — callers should resolve K/L/P before calling this.
        return "L" if holiday_set else "K"

    # row exists → has checkin
    exc_status, exc_type = exception_map.get((row.employee_id, row.work_date), (None, None))
    if exc_type == "MISSED_CHECKOUT":
        return "1/2T"

    punctuality = _rank_to_punctuality(row.punctuality_rank)
    if punctuality == "LATE":
        return "1/2T"

    geo_name = row.checkin_matched_geofence or row.checkout_matched_geofence
    geo_location_type = (
        (geofence_location_map or {}).get((row.group_id, geo_name))
        if geo_name else None
    )
    geo_type = _geofence_type(geo_name, geo_location_type)
    shift_start = row.shift_start or time(8, 0)
    shift_end = row.shift_end or time(17, 0)
    regular_minutes, _, _ = split_regular_overtime_minutes(
        row.work_date,
        row.checkin_time,
        row.checkout_time,
        shift_start,
        shift_end,
    )
    full_day = (regular_minutes or 0) >= 480

    if geo_type == "VP":
        return "V" if full_day else "1/2V"
    if geo_type == "SITE":
        return "S" if full_day else "1/2S"
    # SYSTEM_RULE
    return "X" if full_day else "1/2T"


def _calc_summary(codes: list[str]) -> dict:
    """Compute aggregate columns from a list of cell codes for one employee."""
    return {
        "ngay_cong": sum(
            1.0 if c in _FULL_CODES else 0.5 if c in _HALF_CODES else 0.0
            for c in codes
        ),
        "tai_vp": sum(1.0 if c == "V" else 0.5 if c == "1/2V" else 0.0 for c in codes),
        "tai_site": sum(1.0 if c == "S" else 0.5 if c == "1/2S" else 0.0 for c in codes),
        "vang_k": codes.count("K"),
        "nghi_le": codes.count("L"),
        "nghi_p": codes.count("P"),
    }


# ---------------------------------------------------------------------------
# Monthly Attendance Matrix — Phase 1: endpoint + data pivot
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Monthly Attendance Matrix — Phase 2: Excel builder
# ---------------------------------------------------------------------------

_VN_WEEKDAY = {0: "T2", 1: "T3", 2: "T4", 3: "T5", 4: "T6", 5: "T7", 6: "CN"}

_CODE_FILLS = {
    "V":    PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
    "S":    PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid"),
    "X":    None,
    "1/2V": PatternFill(start_color="FED7AA", end_color="FED7AA", fill_type="solid"),
    "1/2S": PatternFill(start_color="FED7AA", end_color="FED7AA", fill_type="solid"),
    "1/2T": PatternFill(start_color="FED7AA", end_color="FED7AA", fill_type="solid"),
    "K":    PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid"),
    "L":    PatternFill(start_color="E9D5FF", end_color="E9D5FF", fill_type="solid"),
    "P":    PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
}
_SAT_FILL     = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
_SUN_FILL     = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
_HOL_FILL     = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")
_HDR_FILL     = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
_SUBHDR_FILL  = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
_SUM_FILL     = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")

_SUMMARY_HEADERS = ["Ngày công", "Tại VP", "Tại Site", "Nghỉ lễ (L)", "Nghỉ lương (P)", "Vắng (K)"]
_SUMMARY_KEYS    = ["ngay_cong", "tai_vp", "tai_site", "nghi_le", "nghi_p", "vang_k"]
_SUMMARY_WIDTHS  = [10, 9, 9, 11, 13, 9]

_LEGEND_ITEMS = [
    ("V",    "Đủ công ≥ 8h tại Văn phòng"),
    ("S",    "Đủ công ≥ 8h tại Site"),
    ("X",    "Đủ công ≥ 8h (không có geofence)"),
    ("1/2V", "Làm < 8h tại Văn phòng"),
    ("1/2S", "Làm < 8h tại Site"),
    ("1/2T", "Đi trễ hoặc quên chấm công"),
    ("K",    "Vắng không lương"),
    ("L",    "Nghỉ lễ"),
    ("P",    "Nghỉ có lương (chờ module nghỉ phép)"),
]

_CENTER  = Alignment(horizontal="center", vertical="center")
_LEFT    = Alignment(horizontal="left",   vertical="center")
_WRAP    = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _day_col_fill(weekday: int, is_holiday: bool) -> PatternFill | None:
    if is_holiday:
        return _HOL_FILL
    if weekday == 6:
        return _SUN_FILL
    if weekday == 5:
        return _SAT_FILL
    return _SUBHDR_FILL


def _build_monthly_excel(
    *,
    year: int,
    month: int,
    group_id: int | None,
    all_employees: list,
    emp_meta: dict[int, tuple[str, str]],
    days: list[int],
    matrix: dict[int, list[str]],
    summaries: dict[int, dict],
    holiday_set: set,
) -> StreamingResponse:
    last_day = days[-1]

    # Column layout: A=1 STT | B=2 MãNV | C=3 Tên | D..=day cols | then summary cols
    DATA_COL = 4
    SUM_COL  = DATA_COL + last_day
    TOT_COLS = SUM_COL + len(_SUMMARY_HEADERS) - 1

    # Per-day metadata
    day_wd  = [date(year, month, d).weekday() for d in days]   # 0=Mon … 6=Sun
    day_hol = [date(year, month, d) in holiday_set for d in days]

    # --- Employee order: active sorted by code, then extras with logs ---
    active_ids = {emp.id for emp in all_employees}
    emp_order  = [emp.id for emp in all_employees if emp.id in matrix]
    for eid in matrix:
        if eid not in active_ids:
            emp_order.append(eid)

    wb = Workbook()
    ws = wb.active
    ws.title = f"T{month:02d}-{year}"

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOT_COLS)
    c = ws.cell(1, 1)
    c.value     = f"BẢNG CHẤM CÔNG THÁNG {month:02d}/{year}"
    c.font      = Font(bold=True, size=14, color="1E3A8A")
    c.alignment = _CENTER
    ws.row_dimensions[1].height = 28

    # ── Row 2: Column headers ─────────────────────────────────────────────────
    for col, label in ((1, "STT"), (2, "Mã NV"), (3, "Họ và Tên")):
        c = ws.cell(2, col)
        c.value     = label
        c.fill      = _HDR_FILL
        c.font      = Font(bold=True, size=10, color="FFFFFF")
        c.alignment = _CENTER

    for i, d in enumerate(days):
        col = DATA_COL + i
        c = ws.cell(2, col)
        c.value     = f"{d:02d}"
        c.fill      = _day_col_fill(day_wd[i], day_hol[i])
        c.font      = Font(bold=True, size=10,
                           color="7C3AED" if day_hol[i] else
                                 "D97706" if day_wd[i] == 6 else
                                 "6B7280" if day_wd[i] == 5 else "1E3A8A")
        c.alignment = _CENTER

    for i, hdr in enumerate(_SUMMARY_HEADERS):
        c = ws.cell(2, SUM_COL + i)
        c.value     = hdr
        c.fill      = _SUM_FILL
        c.font      = Font(bold=True, size=10, color="065F46")
        c.alignment = _WRAP
    ws.row_dimensions[2].height = 22

    # ── Row 3: Weekday names ──────────────────────────────────────────────────
    for col in (1, 2, 3):
        c = ws.cell(3, col)
        c.fill = _HDR_FILL
        c.font = Font(bold=True, size=9, color="FFFFFF")

    for i, d in enumerate(days):
        col = DATA_COL + i
        c = ws.cell(3, col)
        c.value     = _VN_WEEKDAY[day_wd[i]]
        c.fill      = _day_col_fill(day_wd[i], day_hol[i])
        c.font      = Font(size=9,
                           color="7C3AED" if day_hol[i] else
                                 "D97706" if day_wd[i] == 6 else
                                 "6B7280" if day_wd[i] == 5 else "1E3A8A")
        c.alignment = _CENTER

    for i in range(len(_SUMMARY_HEADERS)):
        ws.cell(3, SUM_COL + i).fill = _SUM_FILL
    ws.row_dimensions[3].height = 15

    # ── Data rows ─────────────────────────────────────────────────────────────
    for seq, eid in enumerate(emp_order, start=1):
        r        = 3 + seq  # row 4, 5, …
        emp_code, full_name = emp_meta.get(eid, ("", ""))
        codes    = matrix[eid]
        summary  = summaries[eid]

        ws.cell(r, 1).value     = seq
        ws.cell(r, 1).alignment = _CENTER
        ws.cell(r, 2).value     = emp_code
        ws.cell(r, 2).alignment = _CENTER
        ws.cell(r, 3).value     = full_name
        ws.cell(r, 3).alignment = _LEFT

        for i, code in enumerate(codes):
            col  = DATA_COL + i
            cell = ws.cell(r, col)
            cell.alignment = _CENTER

            # Background: code fill > weekend col fill
            col_fill = _day_col_fill(day_wd[i], day_hol[i])
            code_fill = _CODE_FILLS.get(code) if code else None
            if code_fill is not None:
                cell.fill = code_fill
            elif not code and col_fill is not None:
                cell.fill = col_fill
            elif col_fill is not None and code in ("", None):
                cell.fill = col_fill

            if not code:
                continue  # weekend blank — no value

            cell.value = code
            if code == "K":
                cell.font = Font(bold=True, size=10, color="DC2626")
            elif code == "L":
                cell.font = Font(bold=True, size=10, color="7C3AED")
            elif code in _FULL_CODES:
                cell.font = Font(bold=True, size=10, color="1D4ED8")
            else:
                cell.font = Font(size=10, color="92400E")

        for i, key in enumerate(_SUMMARY_KEYS):
            col  = SUM_COL + i
            cell = ws.cell(r, col)
            val  = summary.get(key, 0)
            # Display as int if whole number, else float with 1 decimal
            cell.value     = int(val) if val == int(val) else round(val, 1)
            cell.alignment = _CENTER
            cell.fill      = _SUM_FILL
            cell.font      = (Font(bold=True, size=10, color="065F46") if val > 0
                               else Font(size=10, color="9CA3AF"))

        ws.row_dimensions[r].height = 18

    # ── Legend ────────────────────────────────────────────────────────────────
    legend_row = 3 + len(emp_order) + 2
    ws.cell(legend_row, 1).value = "CHÚ THÍCH:"
    ws.cell(legend_row, 1).font  = Font(bold=True, size=10)

    for i, (code, desc) in enumerate(_LEGEND_ITEMS):
        r    = legend_row + 1 + i
        cc   = ws.cell(r, 1)
        cc.value     = code
        cc.alignment = _CENTER
        cc.font      = Font(bold=True, size=10)
        fill = _CODE_FILLS.get(code)
        if fill:
            cc.fill = fill

        dc = ws.cell(r, 2)
        dc.value = desc
        dc.font  = Font(size=10)
        dc.alignment = _LEFT

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 24
    for i in range(last_day):
        ws.column_dimensions[get_column_letter(DATA_COL + i)].width = 4.5
    for i, w in enumerate(_SUMMARY_WIDTHS):
        ws.column_dimensions[get_column_letter(SUM_COL + i)].width = w

    # ── Freeze: rows 1-3 + cols A-C ──────────────────────────────────────────
    ws.freeze_panes = "D4"

    # ── Stream ────────────────────────────────────────────────────────────────
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    group_suffix = f"_nhom_{group_id}" if group_id is not None else ""
    filename = f"cham_cong_thang_{month:02d}_{year}{group_suffix}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/attendance-monthly.xlsx")
def export_monthly_attendance_excel(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020, le=2100),
    group_id: int | None = None,
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    """Export monthly attendance matrix as Excel.

    Returns a spreadsheet with one row per employee and one column per calendar
    day, filled with attendance codes (V/S/X/1/2V/1/2S/1/2T/K/L/P).
    Summary columns (ngày công, tại VP, tại Site, nghỉ lễ, nghỉ P, vắng K) are
    appended at the right.
    """
    import calendar

    from_date = date(year, month, 1)
    last_day = calendar.monthrange(year, month)[1]
    to_date = date(year, month, last_day)

    if group_id is not None:
        group = db.query(Group).filter(Group.id == group_id).first()
        if not group:
            raise HTTPException(status_code=404, detail="group_id không tồn tại")

    # --- Load attendance rows (employees who have checkin records) ---
    rows = _fetch_daily_report_rows(db, from_date, to_date, None, group_id)
    exception_map = _build_exception_map(db, from_date, to_date, None, group_id)

    # --- Load public holidays for the month ---
    from sqlalchemy import extract as sa_extract
    holiday_rows = (
        db.query(PublicHoliday.date)
        .filter(
            sa_extract("year", PublicHoliday.date) == year,
            sa_extract("month", PublicHoliday.date) == month,
        )
        .all()
    )
    holiday_set: set[date] = {r.date for r in holiday_rows}

    # --- Build (group_id, name) → location_type lookup ---
    # Keyed by (group_id, name) to avoid collision when two groups share a geofence name
    # but have different location_type values.
    gf_q = db.query(GroupGeofence.group_id, GroupGeofence.name, GroupGeofence.location_type)
    if group_id is not None:
        gf_q = gf_q.filter(GroupGeofence.group_id == group_id)
    geofence_location_map: dict[tuple[int, str], str] = {
        (gid, gname): lt for gid, gname, lt in gf_q.all()
    }

    # --- Load approved leave requests overlapping this month ---
    leave_map = _build_leave_map(db, from_date, to_date, group_id)

    # --- Load ALL active employees for the period (include those with 0 checkins) ---
    emp_q = db.query(Employee).filter(
        Employee.active.is_(True),
        Employee.deleted_at.is_(None),
    )
    if group_id is not None:
        emp_q = emp_q.filter(Employee.group_id == group_id)
    all_employees = emp_q.order_by(Employee.code.asc()).all()

    # --- Pivot: {employee_id: {work_date: row}} ---
    pivot: dict[int, dict[date, object]] = {emp.id: {} for emp in all_employees}
    emp_meta: dict[int, tuple[str, str]] = {
        emp.id: (emp.code, emp.full_name) for emp in all_employees
    }

    for row in rows:
        eid = int(row.employee_id)
        if eid not in pivot:
            # employee visible via logs but not in current active filter — include anyway
            pivot[eid] = {}
            emp_meta[eid] = (row.employee_code, row.full_name)
        pivot[eid][row.work_date] = row

    # --- Build cell codes matrix ---
    days = list(range(1, last_day + 1))
    matrix: dict[int, list[str]] = {}
    for eid in pivot:
        emp_rows = pivot[eid]
        codes: list[str] = []
        for d in days:
            day_date = date(year, month, d)
            weekday = day_date.weekday()  # 5=Sat, 6=Sun
            if weekday >= 5:
                # Weekend — skip for K counting, but show actual code if worked
                row = emp_rows.get(day_date)
                if row is not None:
                    # Worked on weekend — derive actual code
                    is_holiday = day_date in holiday_set
                    codes.append(_derive_cell_code(row, exception_map, {day_date} if is_holiday else set(), geofence_location_map))
                else:
                    codes.append("")  # weekend, no work
            else:
                row = emp_rows.get(day_date)
                is_holiday = day_date in holiday_set
                if row is None:
                    leave_code = leave_map.get((eid, day_date))
                    if leave_code == "P":
                        # PAID leave always wins — even on public holidays (P beats L)
                        codes.append("P")
                    elif is_holiday:
                        codes.append("L")
                    else:
                        # UNPAID leave or plain absent both map to K
                        codes.append("K")
                else:
                    codes.append(_derive_cell_code(row, exception_map, {day_date} if is_holiday else set(), geofence_location_map))
        matrix[eid] = codes

    # --- Compute summaries ---
    summaries: dict[int, dict] = {
        eid: _calc_summary(codes) for eid, codes in matrix.items()
    }

    return _build_monthly_excel(
        year=year,
        month=month,
        group_id=group_id,
        all_employees=all_employees,
        emp_meta=emp_meta,
        days=days,
        matrix=matrix,
        summaries=summaries,
        holiday_set=holiday_set,
    )
